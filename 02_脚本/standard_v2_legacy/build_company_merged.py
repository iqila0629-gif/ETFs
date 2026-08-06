"""Build company-format merged wide tables from per-signal CSVs."""

from __future__ import annotations

import csv
import os
import pathlib
import sys

import openpyxl
from openpyxl.utils import get_column_letter


def parse_name(stem: str) -> tuple[str, str, int]:
    ticker, rest = stem.split("__", 1)
    horizon = 1
    if "__N" in rest:
        rest, h = rest.split("__N")
        horizon = int(h)
    return ticker, rest, horizon


def read_signals(csv_dir: pathlib.Path):
    headers: list[str] = []
    data: dict[str, dict[str, float]] = {}
    all_dates: set[str] = set()
    for p in sorted(csv_dir.glob("*.csv")):
        ticker, condition, horizon = parse_name(p.stem)
        label = f"{ticker}__{condition}" + (f"__N{horizon}" if horizon > 1 else "")
        headers.append(label)
        mapping: dict[str, float] = {}
        with p.open("r", encoding="utf-8-sig", newline="") as fh:
            for r, row in enumerate(csv.reader(fh)):
                if r <= 12 or len(row) < 2:
                    continue
                s = row[1].strip().replace(",", "")
                if s in ("", "nan"):
                    continue
                mapping[row[0]] = float(s)
                all_dates.add(row[0])
        data[label] = mapping
    return headers, data, all_dates


def write_company_wide(path: pathlib.Path, headers: list[str], data: dict[str, dict[str, float]], all_dates: set[str]) -> None:
    dates = sorted(all_dates, key=lambda d: (int(d[6:]), int(d[0:2]), int(d[3:5])), reverse=True)
    n_data = len(dates)
    last_row = 13 + n_data

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Sheet1")
    ws.append([])
    stat_templates = {
        2: lambda col, letter: f"={letter}3/({letter}3+{letter}4)",
        3: lambda col, letter: f'=COUNTIF({letter}14:{letter}{last_row},">0")',
        4: lambda col, letter: f'=COUNTIF({letter}14:{letter}{last_row},"<0")',
        5: lambda col, letter: f"=AVERAGE({letter}14:{letter}{last_row})",
        6: lambda col, letter: f"=MAX({letter}14:{letter}{last_row})",
        7: lambda col, letter: f"=MIN({letter}14:{letter}{last_row})",
        8: lambda col, letter: f"=COUNT({letter}14:{letter}{last_row})",
        9: lambda col, letter: f"=STDEV({letter}14:{letter}{last_row})",
        10: lambda col, letter: f"=SUM({letter}14:{letter}{last_row})",
    }
    labels = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
    for stat_row, label in zip(range(2, 11), labels):
        out = [label]
        for col in range(2, len(headers) + 2):
            out.append(stat_templates[stat_row](col, get_column_letter(col)))
        ws.append(out)
    ws.append([])
    ws.append([])
    ws.append(["Date"] + headers)
    for d in dates:
        row = [d]
        for label in headers:
            row.append(data[label].get(d, ""))
        ws.append(row)
    wb.save(path)


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    jobs = [
        (pathlib.Path(os.environ["FULL_CSV"]), pathlib.Path(os.environ["FULL_OUT"])),
        (pathlib.Path(os.environ["FROZEN_CSV"]), pathlib.Path(os.environ["FROZEN_OUT"])),
    ]
    for csv_dir, out_path in jobs:
        headers, data, all_dates = read_signals(csv_dir)
        write_company_wide(out_path, headers, data, all_dates)
        print(out_path.name, "signals:", len(headers), "dates:", len(all_dates))


if __name__ == "__main__":
    main()
