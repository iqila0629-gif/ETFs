"""Rename delivery files to Chinese and add 说明 (Sheet1) + 数据 (Sheet2)."""

from __future__ import annotations

import csv
import pathlib
import sys

import openpyxl
from openpyxl.utils import get_column_letter


ROOT = pathlib.Path(r"C:\Users\vanessacen\Desktop\基金预测")
DELIV = ROOT / "交付物_2026-08-03"
SRC = ROOT / "analysis_results" / "standard_v2"
EVENT = ROOT / "analysis_results" / "event_study"
PROC = ROOT / "processed_returns"


def to_number(value):
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace(",", "")
    if s in ("", "nan", "None"):
        return ""
    try:
        return float(s)
    except ValueError:
        return s


def new_wb(explain_lines):
    wb = openpyxl.Workbook(write_only=True)
    ws1 = wb.create_sheet(title="说明")
    for line in explain_lines:
        ws1.append([line])
    ws2 = wb.create_sheet(title="数据")
    return wb, ws2


def build_plain(csv_path, out_path, explain, colmap=None, skip_header=False):
    wb, ws = new_wb(explain)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        if skip_header:
            pass
        if colmap:
            ws.append([colmap.get(h, h) for h in header])
        else:
            ws.append(header)
        for row in reader:
            ws.append([to_number(v) for v in row])
    wb.save(out_path)


def build_company13(csv_path, out_path, explain, name_col="Date"):
    rows = list(csv.reader(csv_path.open("r", encoding="utf-8-sig", newline="")))
    header_idx = 12
    ncols = len(rows[header_idx]) if len(rows) > header_idx else 0
    data = []
    for r, row in enumerate(rows):
        if r > header_idx:
            data.append([v if c == 0 else to_number(v) for c, v in enumerate(row)])
    has_value = [False] * (ncols + 1)
    for row in data:
        for c in range(2, ncols + 1):
            if c - 1 < len(row) and row[c - 1] not in ("", None):
                has_value[c] = True

    wb, ws = new_wb(explain)
    ws.append([])
    last = 13 + len(data)
    stat_templates = {
        2: lambda letter: f"={letter}3/({letter}3+{letter}4)",
        3: lambda letter: f'=COUNTIF({letter}14:{letter}{last},">0")',
        4: lambda letter: f'=COUNTIF({letter}14:{letter}{last},"<0")',
        5: lambda letter: f"=AVERAGE({letter}14:{letter}{last})",
        6: lambda letter: f"=MAX({letter}14:{letter}{last})",
        7: lambda letter: f"=MIN({letter}14:{letter}{last})",
        8: lambda letter: f"=COUNT({letter}14:{letter}{last})",
        9: lambda letter: f"=STDEV({letter}14:{letter}{last})",
        10: lambda letter: f"=SUM({letter}14:{letter}{last})",
    }
    labels = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
    for stat_row, label in zip(range(2, 11), labels):
        out = [label]
        for col in range(2, ncols + 1):
            out.append(stat_templates[stat_row](get_column_letter(col)) if has_value[col] else "")
        ws.append(out)
    ws.append([])
    ws.append([])
    ws.append(rows[header_idx])
    for row in data:
        ws.append(row)
    wb.save(out_path)


def build_wide_merged(csv_dir, out_path, explain):
    headers = []
    data = {}
    all_dates = set()
    for p in sorted(csv_dir.glob("*.csv")):
        stem = p.stem
        ticker, rest = stem.split("__", 1)
        horizon = 1
        if "__N" in rest:
            rest, h = rest.split("__N")
            horizon = int(h)
        label = f"{ticker}__{rest}" + (f"__N{horizon}" if horizon > 1 else "")
        headers.append(label)
        mapping = {}
        for r, row in enumerate(csv.reader(p.open("r", encoding="utf-8-sig", newline=""))):
            if r <= 12 or len(row) < 2:
                continue
            s = row[1].strip().replace(",", "")
            if s in ("", "nan"):
                continue
            mapping[row[0]] = float(s)
            all_dates.add(row[0])
        data[label] = mapping
    dates = sorted(all_dates, key=lambda d: (int(d[6:]), int(d[0:2]), int(d[3:5])), reverse=True)
    last = 13 + len(dates)

    wb, ws = new_wb(explain)
    ws.append([])
    stat_templates = {
        2: lambda letter: f"={letter}3/({letter}3+{letter}4)",
        3: lambda letter: f'=COUNTIF({letter}14:{letter}{last},">0")',
        4: lambda letter: f'=COUNTIF({letter}14:{letter}{last},"<0")',
        5: lambda letter: f"=AVERAGE({letter}14:{letter}{last})",
        6: lambda letter: f"=MAX({letter}14:{letter}{last})",
        7: lambda letter: f"=MIN({letter}14:{letter}{last})",
        8: lambda letter: f"=COUNT({letter}14:{letter}{last})",
        9: lambda letter: f"=STDEV({letter}14:{letter}{last})",
        10: lambda letter: f"=SUM({letter}14:{letter}{last})",
    }
    labels = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
    for stat_row, label in zip(range(2, 11), labels):
        ws.append([label] + [stat_templates[stat_row](get_column_letter(col)) for col in range(2, len(headers) + 2)])
    ws.append([])
    ws.append([])
    ws.append(["Date"] + headers)
    for d in dates:
        ws.append([d] + [data[label].get(d, "") for label in headers])
    wb.save(out_path)


COL_MAP = {
    "ticker": "基金代码",
    "fund_group": "基金类型",
    "source": "信号来源",
    "condition": "条件",
    "horizon": "窗口(日)",
    "full_avg": "全历史平均回报(%)",
    "full_trades": "全历史交易数",
    "full_hit": "全历史命中率",
    "frozen_avg": "冻结期平均回报(%)",
    "frozen_trades": "冻结期交易数",
    "frozen_hit": "冻结期命中率",
    "dual_pass": "是否双口径达标",
    "best_condition": "最佳条件",
    "best_horizon": "最佳窗口(日)",
    "predicted_return": "预测回报(%)",
    "actual_return": "实际回报(%)",
    "date": "日期",
    "Date": "日期",
    "ETF": "ETF代码",
    "signals": "信号次数",
    "funds": "覆盖基金数",
    "tier": "档位",
    "small_sample": "样本小",
}


def explain_signals(title):
    return [
        f"本表：{title}",
        "",
        "数据 Sheet 每行 = 一个信号。",
        "",
        "列说明：",
        "基金代码 = ProFunds 基金，如 UOPIX",
        "基金类型 = long（普通多头）/ ultra_long（杠杆多头）/ inverse（反向）",
        "信号来源 = main（主模型）/ sparse（专项）/ optimization_scan（优化扫描）/ gap_extra（缺口补充）",
        "条件 = 触发条件，如 XLY_lt-2 表示 XLY 当日跌超 2%",
        "窗口(日) = 预测未来 N 日，1 = 次日",
        "全历史平均回报(%) = 全历史出手日次日实际回报的平均值",
        "全历史交易数 = 全历史出手次数",
        "全历史命中率 = 预测方向正确的比例",
        "冻结期平均回报(%) = 2025-2026 出手日实际回报的平均值",
        "冻结期交易数 = 2025-2026 出手次数",
        "冻结期命中率 = 2025-2026 方向正确比例",
        "是否双口径达标 = 全历史 + 冻结期是否都达标",
    ]


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    # root company tables
    company_pairs = [
        (SRC / "v2_company_daily_best_full_history.csv", DELIV / "公司格式_最佳策略_全历史.xlsx",
         ["本表：每支基金最佳策略，全历史。", "", "数据 Sheet = 公司 13 行格式：",
          "第 2-10 行为统计公式，第 13 行为 Date + 基金代码，数据行为每日触发记录。",
          "有值 = 该基金最佳信号触发，值为次日实际回报；空白 = 不操作。"]),
        (SRC / "v2_company_daily_best_frozen.csv", DELIV / "公司格式_最佳策略_冻结期.xlsx",
         ["本表：每支基金最佳策略，冻结期 2025-2026。", "结构同全历史版，只保留冻结期。"]),
        (SRC / "v2_company_daily_all_signals_full_history.csv", DELIV / "公司格式_全信号_全历史.xlsx",
         ["本表：每支基金全部正式信号按 R4 合并，全历史。", "数据 Sheet = 公司 13 行格式；",
          "每列 = 一支基金，值 = 该基金任一信号触发日实际回报，冲突日按最强信号。",
          "注意：合并后基金级平均可能被弱信号稀释，不作为验收。"]),
        (SRC / "v2_company_daily_all_signals_frozen.csv", DELIV / "公司格式_全信号_冻结期.xlsx",
         ["本表：每支基金全部正式信号按 R4 合并，冻结期 2025-2026。", "结构同全历史版。"]),
    ]
    for csv_path, out, explain in company_pairs:
        build_company13(csv_path, out, explain)
        print("built", out.name)

    # root signal tables
    plain_pairs = [
        (SRC / "v2_dual_criteria_pass.csv", DELIV / "正式信号明细_全量.xlsx", explain_signals("v2 双口径正式信号全量明细，1,455 个")),
        (SRC / "v2_dual_criteria_summary.csv", DELIV / "每基金最佳信号汇总.xlsx", explain_signals("每支基金只保留 1 个最佳信号，85 行")),
        (SRC / "v2_strategy_mapping.csv", DELIV / "每基金策略映射.xlsx", explain_signals("每支基金最终采用的策略")),
    ]
    for csv_path, out, explain in plain_pairs:
        build_plain(csv_path, out, explain, COL_MAP)
        print("built", out.name)

    build_plain(
        SRC / "v2_etf_usage.csv",
        DELIV / "ETF使用统计.xlsx",
        ["本表：每支 ETF 在正式信号中的使用统计。", "", "列说明：",
         "ETF代码 = ETF 代码；信号次数 = 出现在多少个正式信号中；覆盖基金数 = 至少一个信号用到它的基金数。"],
        COL_MAP,
    )
    print("built ETF使用统计.xlsx")

    for csv_path, out, title in [
        (SRC / "v2_signal_detail_full_history.csv", DELIV / "信号逐日明细_全历史.xlsx", "全部正式信号逐日记录，全历史"),
        (SRC / "v2_signal_detail_frozen.csv", DELIV / "信号逐日明细_冻结期.xlsx", "全部正式信号逐日记录，冻结期"),
    ]:
        build_plain(
            csv_path,
            out,
            [f"本表：{title}。", "", "数据 Sheet 每行 = 一次触发。",
             "基金代码 = 基金；条件 = 触发条件；窗口(日) = 预测窗口；",
             "日期 = 触发日；预测回报(%) = 历史方向平均回报；实际回报(%) = 次日实际回报。"],
            COL_MAP,
        )
        print("built", out.name)

    # processed merged data (company 13-line)
    data_pairs = [
        (PROC / "combined_profunds_nav.csv", DELIV / "数据_处理后合并" / "合并_基金净值_129支.xlsx",
         ["本表：129 支基金 NAV 合并表（公司 13 行格式）。", "第 13 行为 Date + 基金代码，数据行为每日 NAV。"]),
        (PROC / "combined_etf_returns.csv", DELIV / "数据_处理后合并" / "合并_ETF日回报_原始19支.xlsx",
         ["本表：原始 19 支 ETF 日回报合并表（公司 13 行格式）。"]),
        (PROC / "combined_extended_etf_returns.csv", DELIV / "数据_处理后合并" / "合并_ETF日回报_扩展57支.xlsx",
         ["本表：扩展 57 支 ETF 日回报合并表（公司 13 行格式），2016-08 起。"]),
    ]
    for csv_path, out, explain in data_pairs:
        build_company13(csv_path, out, explain)
        print("built", out.name)

    # panels
    build_plain(EVENT / "panel_fund_returns.csv", DELIV / "数据_处理后合并" / "建模面板_基金日回报.xlsx",
                ["本表：建模用基金日回报面板。", "每行 = 一个交易日；每列 = 一支基金；回报为小数（0.01=1%）。"], {"Date": "日期"})
    build_plain(EVENT / "panel_etf_returns.csv", DELIV / "数据_处理后合并" / "建模面板_ETF日回报.xlsx",
                ["本表：建模用 19 支 ETF 日回报面板。", "每行 = 一个交易日；回报为小数（0.01=1%）。"], {"Date": "日期"})
    print("built panels")

    # merged wide company format
    build_wide_merged(
        SRC / "final_outputs_dual_full_history",
        DELIV / "正式预测_全历史" / "正式预测_全历史_合并.xlsx",
        ["本表：正式预测全历史合并（公司 13 行格式宽表）。",
         "第 13 行 = Date + 1,455 个信号列（基金代码__条件，多日带 __N窗口）。",
         "数据行 = 所有触发日并集；每列只在触发日有值。",
         "第 2-10 行 = 各列统计公式。"],
    )
    build_wide_merged(
        SRC / "final_outputs_dual_frozen",
        DELIV / "正式预测_冻结期" / "正式预测_冻结期_合并.xlsx",
        ["本表：正式预测冻结期合并（公司 13 行格式宽表），结构与全历史版一致。"],
    )
    print("built merged wide tables")

    # uncovered tier tables
    tier_explain = [
        ["本表：未覆盖基金评估。", "每行 = 一支基金最接近目标的信号。"],
    ]
    for csv_path, out in [
        (SRC / "uncovered_funds_eval" / "tier1_frozen_pass.csv", DELIV / "未覆盖基金评估" / "未覆盖_第一档_冻结期已达标.xlsx"),
        (SRC / "uncovered_funds_eval" / "tier2_frozen_pass_small_sample.csv", DELIV / "未覆盖基金评估" / "未覆盖_第二档_冻结期达标样本小.xlsx"),
        (SRC / "uncovered_funds_eval" / "tier3_no_frozen_pass.csv", DELIV / "未覆盖基金评估" / "未覆盖_第三档_冻结期未达标.xlsx"),
    ]:
        build_plain(csv_path, out, tier_explain[0], COL_MAP)
        print("built", out.name)
    build_plain(
        SRC / "uncovered_funds_eval" / "all_candidates_stats.csv",
        DELIV / "未覆盖基金评估" / "未覆盖_全部候选重算.xlsx",
        ["本表：38 支未覆盖基金全部候选信号统一重算结果。", "每行 = 一个候选信号；含义同正式信号表。"],
        COL_MAP,
    )
    print("built uncovered tables")


if __name__ == "__main__":
    main()
