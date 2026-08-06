"""Convert data cells to numbers and merge per-signal prediction files."""

from __future__ import annotations

import os
import pathlib
import sys

import openpyxl


def to_number(value):
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return ""
    s = str(value).strip().replace(",", "")
    if s in ("", "nan", "None"):
        return ""
    try:
        return float(s)
    except ValueError:
        return s


def fix_numeric(path: pathlib.Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    if ws.max_row < 14 or ws.cell(13, 1).value != "Date":
        wb.close()
        return
    for row in range(14, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row, col)
            converted = to_number(cell.value)
            if converted != cell.value:
                cell.value = converted
    wb.save(path)
    wb.close()


def parse_name(stem: str) -> tuple[str, str, int]:
    ticker, rest = stem.split("__", 1)
    horizon = 1
    if "__N" in rest:
        rest, h = rest.split("__N")
        horizon = int(h)
    return ticker, rest, horizon


def merge_folder(folder: pathlib.Path) -> pathlib.Path:
    rows = []
    for p in sorted(folder.glob("*.xlsx")):
        wb = openpyxl.load_workbook(p, read_only=False)
        ws = wb["Sheet1"]
        ticker, condition, horizon = parse_name(p.stem)
        for r in range(14, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            ret_val = ws.cell(r, 2).value
            if date_val is None:
                continue
            ret = to_number(ret_val)
            if ret == "":
                ret = ""
            rows.append(
                {
                    "基金代码": ticker,
                    "条件": condition,
                    "窗口": horizon,
                    "日期": str(date_val),
                    "次日回报(%)": ret,
                }
            )
        wb.close()
    rows.sort(key=lambda x: (x["基金代码"], x["条件"], x["窗口"], x["日期"]), reverse=False)
    rows.sort(key=lambda x: (x["基金代码"], x["条件"], x["窗口"]), reverse=False)
    # stable sort keeps dates descending within a group because rows are read descending
    out_path = folder.parent / (folder.name + "_合并.xlsx")
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="合并数据")
    ws.append(["基金代码", "条件", "窗口", "日期", "次日回报(%)"])
    for row in rows:
        ws.append([row["基金代码"], row["条件"], row["窗口"], row["日期"], row["次日回报(%)"]])
    wb.save(out_path)
    return out_path


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    root = pathlib.Path(os.environ["DELIV_DIR"])

    targets = sorted(root.glob("v2_company_*.xlsx"))
    for folder_name in ["正式预测_全历史", "正式预测_冻结期"]:
        folder = root / folder_name
        if folder.exists():
            targets.extend(sorted(folder.glob("*.xlsx")))
    processed = root / "数据_处理后合并"
    if processed.exists():
        targets.extend(sorted(processed.glob("combined_*.xlsx")))

    for p in targets:
        fix_numeric(p)
    print("numeric fix applied to:", len(targets))

    merged = []
    for folder_name in ["正式预测_全历史", "正式预测_冻结期"]:
        folder = root / folder_name
        out = merge_folder(folder)
        merged.append(out)
        files = list(folder.glob("*.xlsx"))
        for p in files:
            p.unlink()
        print(f"{folder_name}: merged {out.name}, removed {len(files)} files")
    print("merged:", [str(m) for m in merged])


if __name__ == "__main__":
    main()
