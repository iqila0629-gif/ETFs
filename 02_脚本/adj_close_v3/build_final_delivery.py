"""Build the final delivery folder '最新成果' from the delivery preview."""

from __future__ import annotations

import csv
import pathlib
import shutil

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter


ROOT = pathlib.Path(__file__).resolve().parents[2]
SRC = ROOT / "analysis_results" / "adj_close_v3" / "delivery_preview"
DST = ROOT / "最新成果"


def to_number(value: str):
    s = str(value).strip().replace(",", "")
    if s in ("", "nan", "None"):
        return ""
    try:
        return float(s)
    except ValueError:
        return s


def load_name_map() -> dict[str, str]:
    df = pd.read_csv(SRC / "数据" / "基金名称映射.csv")
    return dict(zip(df["ticker"].astype(str), df["name"].astype(str)))


NAME_MAP = load_name_map()


def fund_label(ticker: str) -> str:
    name = NAME_MAP.get(str(ticker), str(ticker))
    return f"{name}（{ticker}）"


def new_wb(explain_lines: list[str]):
    wb = openpyxl.Workbook(write_only=True)
    ws1 = wb.create_sheet(title="说明")
    for line in explain_lines:
        ws1.append([line])
    ws2 = wb.create_sheet(title="数据")
    return wb, ws2


def write_plain(
    csv_path: pathlib.Path,
    out_path: pathlib.Path,
    explain_lines: list[str],
    colmap: dict[str, str],
    drop_cols: list[str] | None = None,
    fund_cols: list[str] | None = None,
) -> None:
    drop = set(drop_cols or [])
    funds = set(fund_cols or [])
    wb, ws = new_wb(explain_lines)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        ws.append([colmap.get(h, h) for h in header if h not in drop])
        for row in reader:
            out = []
            for h, v in zip(header, row):
                if h in drop:
                    continue
                if h in funds:
                    out.append(fund_label(v))
                else:
                    out.append(to_number(v))
            ws.append(out)
    wb.save(out_path)


def write_company(
    csv_path: pathlib.Path,
    out_path: pathlib.Path,
    explain_lines: list[str],
    display_funds: bool = True,
) -> None:
    rows = list(csv.reader(csv_path.open("r", encoding="utf-8-sig", newline="")))
    header_idx = 12
    ncols = len(rows[header_idx])
    last = len(rows)
    has_value = [False] * (ncols + 1)
    for row in rows[header_idx + 1 :]:
        for c in range(2, ncols + 1):
            if c - 1 < len(row) and str(row[c - 1]).strip() not in ("", "nan"):
                has_value[c] = True

    wb, ws = new_wb(explain_lines)
    ws.append([])
    stat_templates = {
        2: lambda letter: f"={letter}3/({letter}3+{letter}4)",
        3: lambda letter: f'=COUNTIF({letter}14:{letter}{last},">0")',
        4: lambda letter: f'=COUNTIF({letter}14:{letter}{last},"<0")',
        5: lambda letter: f"=AVERAGE({letter}14:{letter}{last})",
        6: lambda letter: f"=MAX({letter}14:{letter}{last})",
        7: lambda letter: f"=MIN({letter}14:{letter}{last})",
        8: lambda letter: f"=COUNT({letter}14:{letter}{last})",
        9: lambda letter: f"=STDEV({letter}14:{letter}{last})",
        10: lambda letter: f"=SUM({letter}14:{letter}{last})",
    }
    labels = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
    for stat_row, label in zip(range(2, 11), labels):
        out = [label]
        for col in range(2, ncols + 1):
            out.append(stat_templates[stat_row](get_column_letter(col)) if has_value[col] else "")
        ws.append(out)
    ws.append([])
    ws.append([])
    header = rows[header_idx]
    out_header = ["日期"]
    for x in header[1:]:
        out_header.append(fund_label(x) if display_funds else x)
    ws.append(out_header)
    for row in rows[header_idx + 1 :]:
        out = [row[0]]
        for c, v in enumerate(row[1:], start=2):
            out.append(to_number(v))
        ws.append(out)
    wb.save(out_path)


def explain_signal(title: str) -> list[str]:
    return [
        f"本表：{title}。",
        "",
        "数据Sheet每行 = 一个正式信号。",
        "",
        "列说明：",
        "基金名称（代码） = ProFunds 基金",
        "基金类型 = long / ultra_long / inverse",
        "触发条件 = ETF或外部条件，例如 A涨且B跌",
        "预测窗口（日） = 1 = 次日，2/3 = 未来N日日平均",
        "全历史平均回报（%） = 全历史触发日实际回报平均值",
        "全历史交易数 = 全历史触发次数",
        "全历史命中率 = 方向正确比例",
        "冻结期平均回报（%） = 2025-2026 实际回报平均值",
        "冻结期交易数 = 2025-2026 触发次数",
        "冻结期命中率 = 2025-2026 方向正确比例",
        "",
        "条件命名：",
        "ETF_up = ETF 当日涨",
        "ETF_down = ETF 当日跌",
        "ETF_gt2 / ETF_lt-2 = ETF 当日涨超2% / 跌超2%",
        "A_up_B_down = A 涨且 B 跌",
        "ext_vix_chg_ge5 = VIX 当日变化不低于5",
        "self_3down = 基金自身连续三日下跌",
    ]


def explain_best() -> list[str]:
    return [
        "本表：每支基金最终采用的最佳策略。",
        "",
        "数据Sheet每行 = 一支基金。",
        "列说明：",
        "基金名称（代码） = ProFunds 基金",
        "触发条件 = 最终采用的触发条件",
        "预测窗口（日） = 预测未来N日",
        "全历史/冻结期指标含义与正式信号明细一致。",
        "预测方向 = predict_up / predict_down",
        "",
        "选择优先级：",
        "1. |全历史平均回报| 最大",
        "2. 冻结期命中率最高",
        "3. 全历史命中率最高",
    ]


def explain_company(title: str) -> list[str]:
    return [
        f"本表：{title}。",
        "",
        "数据Sheet = 公司13行格式：",
        "第2-10行为统计公式；第13行为日期+基金；数据行从第14行开始。",
        "每列一支基金，值为该基金策略触发日的实际回报（%），未触发留空。",
        "统计行含义：Hit Ratio=命中率，Up/Down Count=涨/跌次数，",
        "Average=平均回报，Max/Min=最大/最小，Count=交易数，Std=标准差，Sum=合计。",
    ]


def make_dirs() -> None:
    for folder in ["全部ETF", "精简ETF", "原始ETF", "中间文档", "数据"]:
        (DST / folder).mkdir(parents=True, exist_ok=True)
    for version in ["全部ETF", "精简ETF", "原始ETF"]:
        (DST / version / "成果").mkdir(parents=True, exist_ok=True)


COMMON_COLMAP = {
    "ticker": "基金名称（代码）",
    "fund_group": "基金类型",
    "condition": "触发条件",
    "horizon": "预测窗口（日）",
    "full_avg": "全历史平均回报（%）",
    "full_trades": "全历史交易数",
    "full_hit": "全历史命中率",
    "frozen_avg": "冻结期平均回报（%）",
    "frozen_trades": "冻结期交易数",
    "frozen_hit": "冻结期命中率",
    "date": "日期",
    "predicted_return": "预测回报（%）",
    "actual_return": "实际回报（%）",
    "ETF": "ETF代码",
    "in_best": "是否进入最佳策略",
    "best_signals": "最佳策略信号数",
    "best_funds": "最佳策略覆盖基金数",
    "all_signals": "全部信号数",
    "all_funds": "全部信号覆盖基金数",
    "covered": "是否覆盖",
    "reason": "未覆盖原因",
    "decision": "预测方向",
    "etf_count": "ETF数量",
    "covered_funds": "覆盖基金数",
    "available_signals": "可用信号数",
    "uncovered_funds": "未覆盖基金数",
    "median_full_hit": "全历史命中率中位数",
    "min_full_hit": "全历史命中率最低值",
    "median_frozen_hit": "冻结期命中率中位数",
    "min_frozen_hit": "冻结期命中率最低值",
    "median_full_avg_abs": "全历史Average中位数",
    "median_frozen_avg_abs": "冻结期Average中位数",
    "median_frozen_trades": "冻结期交易数中位数",
    "avg_min": "两期Average中位数较小值",
    "etf_list": "ETF列表",
    "strict_pass": "是否正式口径通过",
    "frozen_pass": "是否冻结期通过",
    "tier": "档位",
    "small_sample": "样本小",
}


def build_signals(src_csv: pathlib.Path, out: pathlib.Path, title: str) -> None:
    write_plain(
        src_csv,
        out,
        explain_signal(title),
        COMMON_COLMAP,
        drop_cols=["source", "abs_full_avg"],
        fund_cols=["ticker"],
    )


def build_best(src_csv: pathlib.Path, out: pathlib.Path) -> None:
    write_plain(
        src_csv,
        out,
        explain_best(),
        COMMON_COLMAP,
        drop_cols=["source", "abs_full_avg"],
        fund_cols=["ticker"],
    )


def main() -> None:
    make_dirs()

    all_out = DST / "全部ETF" / "成果"
    all_src = SRC / "全部ETF" / "成果"
    build_signals(all_src / "v3_正式信号_全量.csv", all_out / "正式信号明细_全量.xlsx", "全部ETF正式信号")
    build_best(all_src / "v3_每基金最佳策略.csv", all_out / "每基金策略映射.xlsx")
    write_company(
        all_src / "v3_公司格式_最佳策略_全历史.csv",
        all_out / "公司格式_最佳策略_全历史.xlsx",
        explain_company("全部ETF每基金最佳策略，全历史"),
    )
    write_company(
        all_src / "v3_公司格式_最佳策略_冻结期.csv",
        all_out / "公司格式_最佳策略_冻结期.xlsx",
        explain_company("全部ETF每基金最佳策略，冻结期2025-2026"),
    )
    write_company(
        all_src / "v3_公司格式_全信号_全历史.csv",
        all_out / "公司格式_全信号_全历史.xlsx",
        explain_company("全部ETF全信号R4合并，全历史"),
    )
    write_company(
        all_src / "v3_公司格式_全信号_冻结期.csv",
        all_out / "公司格式_全信号_冻结期.xlsx",
        explain_company("全部ETF全信号R4合并，冻结期2025-2026"),
    )
    write_plain(
        all_src / "v3_信号逐日明细_最佳策略.csv",
        all_out / "信号逐日明细_最佳策略.xlsx",
        ["本表：全部ETF最佳策略逐日触发记录。", "", "每行 = 一次触发。", "列含义见正式信号明细说明。"],
        COMMON_COLMAP,
        fund_cols=["ticker"],
    )
    write_plain(
        all_src / "v3_ETF使用统计.csv",
        all_out / "ETF使用统计.xlsx",
        ["本表：每支ETF在正式信号与最佳策略中的使用情况。", "", "每行 = 一支ETF。"],
        COMMON_COLMAP,
    )
    write_plain(
        all_src / "v3_覆盖情况.csv",
        all_out / "覆盖情况.xlsx",
        ["本表：129支基金逐支覆盖状态。", "", "reason=covered/money_fund/not_qualified。"],
        COMMON_COLMAP,
        fund_cols=["ticker"],
    )

    unc = all_out / "未覆盖基金评估"
    unc.mkdir(parents=True, exist_ok=True)
    for src_name, out_name, title in [
        ("v3_未覆盖_全部候选重算.csv", "未覆盖_全部候选重算.xlsx", "未覆盖基金全部候选重算"),
        ("v3_未覆盖_第一档_冻结期已达标.csv", "未覆盖_第一档_冻结期已达标.xlsx", "未覆盖基金第一档"),
        ("v3_未覆盖_第二档_冻结期达标样本小.csv", "未覆盖_第二档_冻结期达标样本小.xlsx", "未覆盖基金第二档"),
        ("v3_未覆盖_第三档_冻结期未达标.csv", "未覆盖_第三档_冻结期未达标.xlsx", "未覆盖基金第三档"),
    ]:
        write_plain(
            all_src / "未覆盖基金评估" / src_name,
            unc / out_name,
            ["本表：" + title + "。", "", "每行 = 一个候选信号；基金按基金名称（代码）展示。"],
            COMMON_COLMAP,
            drop_cols=["dual_pass", "abs_full_avg", "abs_frozen_avg", "source"],
            fund_cols=["ticker"],
        )
    shutil.copy2(all_src / "未覆盖基金评估" / "情况说明.txt", unc / "情况说明.txt")

    stream_src = SRC / "精简ETF" / "成果"
    stream_out = DST / "精简ETF" / "成果"
    build_signals(stream_src / "正式信号_全量.csv", stream_out / "正式信号明细_全量.xlsx", "精简ETF正式信号")
    build_best(stream_src / "每基金最佳策略.csv", stream_out / "每基金策略映射.xlsx")
    write_company(
        stream_src / "公司格式_最佳策略_全历史.csv",
        stream_out / "公司格式_最佳策略_全历史.xlsx",
        explain_company("精简ETF每基金最佳策略，全历史"),
    )
    write_company(
        stream_src / "公司格式_最佳策略_冻结期.csv",
        stream_out / "公司格式_最佳策略_冻结期.xlsx",
        explain_company("精简ETF每基金最佳策略，冻结期2025-2026"),
    )
    write_plain(
        stream_src / "信号逐日明细_最佳策略.csv",
        stream_out / "信号逐日明细_最佳策略.xlsx",
        ["本表：精简ETF最佳策略逐日触发记录。", "", "每行 = 一次触发。"],
        COMMON_COLMAP,
        fund_cols=["ticker"],
    )
    write_plain(
        stream_src / "ETF列表.csv",
        stream_out / "ETF列表.xlsx",
        ["本表：精简版选用的ETF列表。", "", "每行 = 一支ETF。"],
        {"ETF": "ETF代码"},
    )
    write_plain(
        stream_src / "ETF使用统计.csv",
        stream_out / "ETF使用统计.xlsx",
        ["本表：精简版每支ETF使用情况。", "", "每行 = 一支ETF。"],
        COMMON_COLMAP,
    )
    write_plain(
        stream_src / "覆盖情况.csv",
        stream_out / "覆盖情况.xlsx",
        ["本表：129支基金在精简版中的覆盖状态。"],
        COMMON_COLMAP,
        fund_cols=["ticker"],
    )
    write_plain(
        stream_src / "ETF规模影响_可选清单.csv",
        stream_out / "ETF规模影响_可选清单.xlsx",
        ["本表：ETF数量对命中率与Average的影响。", "", "每行 = 一个可选ETF规模。"],
        COMMON_COLMAP,
        drop_cols=["subset_bits", "best_tickers", "best_conditions", "best_horizons", "uncovered_funds"],
    )
    shutil.copy2(stream_src / "精简报告.txt", stream_out / "精简报告.txt")

    orig_src = SRC / "原始ETF" / "成果"
    orig_out = DST / "原始ETF" / "成果"
    build_signals(orig_src / "正式信号_全量.csv", orig_out / "正式信号明细_全量.xlsx", "原始ETF正式信号")
    build_best(orig_src / "每基金最佳策略.csv", orig_out / "每基金策略映射.xlsx")
    write_company(
        orig_src / "公司格式_最佳策略_全历史.csv",
        orig_out / "公司格式_最佳策略_全历史.xlsx",
        explain_company("原始ETF每基金最佳策略，全历史"),
    )
    write_company(
        orig_src / "公司格式_最佳策略_冻结期.csv",
        orig_out / "公司格式_最佳策略_冻结期.xlsx",
        explain_company("原始ETF每基金最佳策略，冻结期2025-2026"),
    )
    write_plain(
        orig_src / "信号逐日明细_最佳策略.csv",
        orig_out / "信号逐日明细_最佳策略.xlsx",
        ["本表：原始ETF最佳策略逐日触发记录。", "", "每行 = 一次触发。"],
        COMMON_COLMAP,
        fund_cols=["ticker"],
    )
    write_plain(
        orig_src / "ETF使用统计.csv",
        orig_out / "ETF使用统计.xlsx",
        ["本表：原始19支ETF使用情况。", "", "每行 = 一支ETF。"],
        COMMON_COLMAP,
    )
    write_plain(
        orig_src / "覆盖情况.csv",
        orig_out / "覆盖情况.xlsx",
        ["本表：129支基金在原始ETF版本中的覆盖状态。"],
        COMMON_COLMAP,
        fund_cols=["ticker"],
    )

    for src in (SRC / "全部ETF" / "说明.txt", SRC / "精简ETF" / "说明.txt", SRC / "精简ETF" / "影响说明.txt", SRC / "原始ETF" / "说明.txt"):
        shutil.copy2(src, DST / src.relative_to(SRC))

    print("key results converted")


if __name__ == "__main__":
    main()
