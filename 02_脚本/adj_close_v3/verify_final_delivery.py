"""Verify 最新成果 final delivery files."""

from __future__ import annotations

import pathlib
import sys

import openpyxl


ROOT = pathlib.Path(__file__).resolve().parents[2]
DST = ROOT / "最新成果"


def check_company(path: pathlib.Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    assert wb.sheetnames[0] == "说明"
    assert wb.sheetnames[1] == "数据"
    ws = wb["数据"]
    rows = list(ws.iter_rows(min_row=2, max_row=10, max_col=3, values_only=True))
    has_formula = any(isinstance(v, str) and v.startswith("=") for row in rows for v in row if v)
    header = list(ws.iter_rows(min_row=13, max_row=13, values_only=True))[0]
    print(path.name, "sheets ok", "formula", has_formula, "header", header[:3])
    wb.close()


def check_plain(path: pathlib.Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    assert wb.sheetnames[0] == "说明"
    assert wb.sheetnames[1] == "数据"
    ws = wb["数据"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    first = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    print(path.name, "header", header[:6], "first", first[:3] if first else None)
    wb.close()


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    company = [
        DST / "全部ETF" / "成果" / "公司格式_最佳策略_全历史.xlsx",
        DST / "全部ETF" / "成果" / "公司格式_最佳策略_冻结期.xlsx",
        DST / "精简ETF" / "成果" / "公司格式_最佳策略_全历史.xlsx",
        DST / "原始ETF" / "成果" / "公司格式_最佳策略_全历史.xlsx",
    ]
    for p in company:
        check_company(p)

    plain = [
        DST / "全部ETF" / "成果" / "每基金策略映射.xlsx",
        DST / "精简ETF" / "成果" / "每基金策略映射.xlsx",
        DST / "原始ETF" / "成果" / "每基金策略映射.xlsx",
        DST / "精简ETF" / "成果" / "ETF列表.xlsx",
        DST / "精简ETF" / "成果" / "ETF规模影响_可选清单.xlsx",
    ]
    for p in plain:
        check_plain(p)

    for version in ["全部ETF", "精简ETF", "原始ETF"]:
        folder = DST / version / "成果"
        files = sorted(p.name for p in folder.iterdir() if p.is_file())
        print(version, files)

    for p in [
        DST / "数据" / "数据_处理后合并" / "合并_基金Adj日回报.xlsx",
        DST / "数据" / "数据_处理后合并" / "合并_ETF日回报_扩展57支.xlsx",
    ]:
        check_company(p)

    for p in [
        DST / "中间文档" / "通用" / "脚本" / "verify_v3_delivery.py",
        DST / "中间文档" / "精简ETF特有" / "etf_streamlined_options.csv",
        DST / "中间文档" / "原始ETF特有" / "original19_pair_strict_pass.csv",
    ]:
        print("intermediate exists", p.name, p.exists())


if __name__ == "__main__":
    main()
