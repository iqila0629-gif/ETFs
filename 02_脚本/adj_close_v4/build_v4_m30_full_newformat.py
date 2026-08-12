"""Build final m30 new-format workbooks for all funds/ETFs (full and frozen)."""

from __future__ import annotations

import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import scan_v4_conditions as sc
import scan_v4_thresholds as s4
import build_v4_m30_newformat_sample as sample


CUTOFF = np.datetime64("2025-01-01")
STAT_NAMES = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
EXT_RAW_FIELDS = {
    "VIX": {"open": "VIX_Open", "high": "VIX_High", "low": "VIX_Low", "close": "VIX_Close"},
    "TNX": {"open": "TNX_Open", "high": "TNX_High", "low": "TNX_Low", "close": "TNX_Yield"},
}
EXT_CHG_NAMES = ["VIX回报", "VIX_Chg%", "VIX_5dChg", "VIX_20dVol", "TNX回报", "TNX_ChgBp"]
EXT_DERIVED = {
    "CreditSpread": "HYG-TLT",
    "JNKSpread": "JNK-TLT",
    "StkBonCorr": "CORREL(SPY,TLT,近20日)",
    "USDGoldRatio": "UUP-GLD",
    "SectRotation": "XLK-XLF",
    "VIX_TNX_Ratio": "VIX_close/TNX_close",
    "YldCurveProxy": "TLT-TIP",
}
EXT_LOGICAL = [
    "VIX_Close", "VIX_Chg%", "TNX_Yield", "TNX_ChgBp", "VIX_5dChg", "VIX_20dVol",
    "CreditSpread", "JNKSpread", "StkBonCorr", "USDGoldRatio", "SectRotation",
    "VIX_TNX_Ratio", "YldCurveProxy",
]

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")
head_font = Font(name="Arial", size=9, bold=True)
body_font = Font(name="Arial", size=9)


def load_mapping() -> pd.DataFrame:
    path = config.V4_OUT / "v4_strategy_mapping_m30_v2.csv"
    return pd.read_csv(path, keep_default_na=False)


def target_formula(fund_col: str, r: int, horizon: int) -> str:
    if horizon == 1:
        return f"{fund_col}{r - 1}"
    return f'IFERROR(ROUND(AVERAGE({fund_col}{r - 1}:{fund_col}{r - horizon}),4),"")'


def build_workbook(
    out_path: str,
    dates_desc: list[pd.Timestamp],
    mapping: pd.DataFrame,
    master: pd.DataFrame,
    title: str,
) -> None:
    name_map = pd.read_csv(config.MIDDLE / "通用" / "文件" / "基金名称映射.csv", keep_default_na=False)
    name_by_ticker = dict(zip(name_map["ticker"], name_map["name"]))
    fund_order = sorted(mapping["ticker"].unique())
    fund_labels = {t: f"{name_by_ticker.get(t, t)}（{t}）" for t in fund_order}
    per_fund = {
        t: [
            (str(r.condition), int(r.horizon))
            for r in mapping[mapping["ticker"] == t].sort_values("strategy_no").itertuples(index=False)
        ]
        for t in fund_order
    }

    etf_all = sorted(config.V4_UNIVERSE)
    ext_cols = list(EXT_LOGICAL)

    base_headers = []
    for t in fund_order:
        base_headers += [f"{fund_labels[t]} Adj Close"]
    for t in fund_order:
        base_headers += [f"{fund_labels[t]}回报"]
    base_headers += [""]
    for e in etf_all:
        base_headers += ["Open", "High", "Low", "Close", "Adj Close", "Volume", ""]
    for e in etf_all:
        base_headers += [f"{e}回报"]
    base_headers += [""]
    base_headers += ["Open", "High", "Low", "Close", ""]
    base_headers += ["Open", "High", "Low", "Close", ""]
    base_headers += EXT_CHG_NAMES + [""]
    base_headers += list(EXT_DERIVED.values()) + [""]

    col_idx = 2
    blank_cols: set[int] = set()
    fund_price_col = {}
    fund_return_col = {}
    for t in fund_order:
        fund_price_col[t] = get_column_letter(col_idx); col_idx += 1
    for t in fund_order:
        fund_return_col[t] = get_column_letter(col_idx); col_idx += 1
    blank_cols.add(col_idx); col_idx += 1

    etf_open_col, etf_high_col, etf_low_col = {}, {}, {}
    etf_close_col, etf_adj_col, etf_volume_col, etf_return_col = {}, {}, {}, {}
    etf_block_start = {}
    for e in etf_all:
        etf_block_start[e] = col_idx
        etf_open_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_high_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_low_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_close_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_adj_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_volume_col[e] = get_column_letter(col_idx); col_idx += 1
        blank_cols.add(col_idx); col_idx += 1
    for e in etf_all:
        etf_return_col[e] = get_column_letter(col_idx); col_idx += 1
    blank_cols.add(col_idx); col_idx += 1

    vix_idx: dict[str, int] = {}
    tnx_idx: dict[str, int] = {}
    vix_start = col_idx
    for key in ("open", "high", "low", "close"):
        vix_idx[key] = col_idx
        col_idx += 1
    blank_cols.add(col_idx); col_idx += 1
    tnx_start = col_idx
    for key in ("open", "high", "low", "close"):
        tnx_idx[key] = col_idx
        col_idx += 1
    blank_cols.add(col_idx); col_idx += 1

    chg_idx: dict[str, int] = {}
    for name in EXT_CHG_NAMES:
        chg_idx[name] = col_idx
        col_idx += 1
    blank_cols.add(col_idx); col_idx += 1

    derived_idx: dict[str, int] = {}
    for name in EXT_DERIVED:
        derived_idx[name] = col_idx
        col_idx += 1
    blank_cols.add(col_idx); col_idx += 1

    ext_col_map = {
        "VIX_Close": get_column_letter(vix_idx["close"]),
        "VIX_Chg%": get_column_letter(chg_idx["VIX_Chg%"]),
        "TNX_Yield": get_column_letter(tnx_idx["close"]),
        "TNX_ChgBp": get_column_letter(chg_idx["TNX_ChgBp"]),
        "VIX_5dChg": get_column_letter(chg_idx["VIX_5dChg"]),
        "VIX_20dVol": get_column_letter(chg_idx["VIX_20dVol"]),
        "CreditSpread": get_column_letter(derived_idx["CreditSpread"]),
        "JNKSpread": get_column_letter(derived_idx["JNKSpread"]),
        "StkBonCorr": get_column_letter(derived_idx["StkBonCorr"]),
        "USDGoldRatio": get_column_letter(derived_idx["USDGoldRatio"]),
        "SectRotation": get_column_letter(derived_idx["SectRotation"]),
        "VIX_TNX_Ratio": get_column_letter(derived_idx["VIX_TNX_Ratio"]),
        "YldCurveProxy": get_column_letter(derived_idx["YldCurveProxy"]),
    }

    colmap = {**fund_return_col, **etf_return_col, **ext_col_map}

    headers = ["日期"] + base_headers
    strat_cols = {}
    merged_idx = {}
    for t in fund_order:
        cols = []
        for cond, _ in per_fund[t]:
            cols.append(col_idx)
            headers.append(sample.condition_label(cond))
            col_idx += 1
        strat_cols[t] = cols
        merged_idx[t] = col_idx
        headers.append("合并效果")
        col_idx += 1
    n_cols = len(headers)

    wb = Workbook()
    ws_note = wb.active
    ws_note.title = "说明"
    ws_note["A1"] = title
    ws_note["A1"].font = Font(name="Arial", size=14, bold=True)
    note_lines = [
        "公式说明：",
        "1. 统计头公式（第2-10行）",
        "   Hit Ratio = Up Count / (Up Count + Down Count)；",
        "   Up Count = COUNTIF(该列数据区, \">0\")；Down Count = COUNTIF(该列数据区, \"<0\")；",
        "   Average = AVERAGE(该列数据区)；Max/Min/Count/Std/Sum 对应 MAX/MIN/COUNT/STDEV/SUM。",
        "2. 回报率公式",
        "   =IF(COUNT(B16:B17)=2,ROUND((B16/B17-1)*100,4),\"\")",
        "   B16 是今日 Adj Close，B17 是前一交易日 Adj Close；",
        "   COUNT 判断两天价格都存在才计算当日回报率，否则留空，避免 #DIV/0!。",
        "   VIX/TNX 原始数据为 Open/High/Low/Close，回报率由 Close 计算；VIX_Chg%/TNX_ChgBp/VIX_5dChg/VIX_20dVol 为变化指标列。",
        "   ETF 原始数据列为 Open/High/Low/Close/Adj Close/Volume，回报率由 Adj Close 计算；",
        "   第14行为区块名称行（如EEM），第15行为列头；ETF列头不重复ETF名。",
        "   派生外部列第14行为名称，第15行为计算公式（谁减谁），数据区为 Excel 公式，如 CreditSpread=HYG回报-TLT回报。",
        "   表内按区块排列：基金区、ETF原始区、ETF回报区、VIX/TNX区、变化区、派生区结束后均用空白列分隔。",
        "   所有回报率均以百分比显示并保留4位小数（公式 ROUND 到4位，显示格式 0.0000%）。",
        "3. 策略公式",
        "   =IF(条件, 该基金次日实际回报, \"\")",
        "   条件引用回报率列和阈值参数；满足条件时显示次日实际回报（日期降序，次日=上一行）。",
        "   horizon>1 时取未来 N 日平均回报，目标公式带 IFERROR，基金无历史数据时留空。",
        "4. 合并公式",
        "   策略列按 |全历史Average| 从高到低排列，合并列取第一个非空策略：",
        "   =IF(策略1<>\"\",策略1,IF(策略2<>\"\",策略2,策略3))",
        "   同一天多条策略触发时，历史 |Average| 最大的策略生效；全部未触发则留空。",
        "5. 阈值参数",
        "   每个策略列正上方的第11/12/13行是该策略条件阈值：二条件用12/13行，三条件额外用11行；",
        "   修改数字可调整条件，公式自动更新。",
    ]
    for i, line in enumerate(note_lines, start=3):
        ws_note.cell(row=i, column=1, value=line)
    ws_note.column_dimensions["A"].width = 120

    ws = wb.create_sheet("数据")
    ws.append([])
    for i, label in enumerate(STAT_NAMES, start=2):
        ws.cell(row=i, column=1, value=label)
    for _ in range(4):
        ws.append([])
    ws.append(headers)
    ds = 16
    de = ds + len(dates_desc) - 1

    for c_idx in range(2, n_cols + 1):
        if c_idx in blank_cols:
            continue
        col = get_column_letter(c_idx)
        ws[f"{col}2"] = f"={col}3/({col}3+{col}4)"
        ws[f"{col}3"] = f'=COUNTIF({col}{ds}:{col}{de},">0")'
        ws[f"{col}4"] = f'=COUNTIF({col}{ds}:{col}{de},"<0")'
        ws[f"{col}5"] = f"=AVERAGE({col}{ds}:{col}{de})"
        ws[f"{col}6"] = f"=MAX({col}{ds}:{col}{de})"
        ws[f"{col}7"] = f"=MIN({col}{ds}:{col}{de})"
        ws[f"{col}8"] = f"=COUNT({col}{ds}:{col}{de})"
        ws[f"{col}9"] = f"=STDEV({col}{ds}:{col}{de})"
        ws[f"{col}10"] = f"=SUM({col}{ds}:{col}{de})"

    fund_adj = {t: sample.load_fund_adj(t) for t in fund_order}
    etf_ohlc = {e: sample.load_etf_ohlc(e) for e in etf_all}
    ext_raw = pd.read_csv(config.EXTERNAL_DAILY, parse_dates=["Date"]).set_index("Date")

    strategy_thr_refs: dict[int, list[str]] = {}
    for t in fund_order:
        for c_idx, (cond, _) in zip(strat_cols[t], per_fund[t]):
            parts = sample.condition_parts(cond, colmap)
            values = [sample.part_threshold(p) for p in parts]
            col = get_column_letter(c_idx)
            ws.cell(row=12, column=c_idx, value=values[0])
            ws.cell(row=13, column=c_idx, value=values[1] if len(values) > 1 else "")
            if len(values) > 2:
                ws.cell(row=11, column=c_idx, value=values[2])
            for r in (11, 12, 13):
                ws.cell(row=r, column=c_idx).font = body_font
            strategy_thr_refs[c_idx] = [f"${col}$12", f"${col}$13"]
            if len(values) > 2:
                strategy_thr_refs[c_idx].append(f"${col}$11")

    for e in etf_all:
        ws.cell(row=14, column=etf_block_start[e] + 2, value=e)
    ws.cell(row=14, column=vix_start + 1, value="VIX")
    ws.cell(row=14, column=tnx_start + 1, value="TNX")
    for name in ("VIX_Chg%", "VIX_5dChg", "VIX_20dVol", "TNX_ChgBp"):
        ws.cell(row=14, column=chg_idx[name], value=name)
    for name in EXT_DERIVED:
        ws.cell(row=14, column=derived_idx[name], value=name)
    for t in fund_order:
        for c_idx in strat_cols[t] + [merged_idx[t]]:
            ws.cell(row=14, column=c_idx, value=fund_labels[t])

    price_cols = (
        [fund_price_col[t] for t in fund_order]
        + [etf_adj_col[e] for e in etf_all]
    )
    ret_cols = (
        [fund_return_col[t] for t in fund_order]
        + [etf_return_col[e] for e in etf_all]
    )

    for i, d in enumerate(dates_desc):
        r = ds + i
        row = [d.strftime("%Y/%m/%d")]
        for t in fund_order:
            row.append(round(fund_adj[t].get(d, float("nan")), 4))
        for t in fund_order:
            row.append("")
        row.append("")
        for e in etf_all:
            rec = etf_ohlc[e].get(d, {})
            row.append(round(rec.get("open", float("nan")), 4))
            row.append(round(rec.get("high", float("nan")), 4))
            row.append(round(rec.get("low", float("nan")), 4))
            row.append(round(rec.get("close", float("nan")), 4))
            row.append(round(rec.get("adj", float("nan")), 4))
            row.append(rec.get("volume", float("nan")))
            row.append("")
        for e in etf_all:
            row.append("")
        row.append("")
        ext_rec = ext_raw.loc[d] if d in ext_raw.index else {}
        for name, idx_map in (("VIX", vix_idx), ("TNX", tnx_idx)):
            for key in ("open", "high", "low", "close"):
                row.append(round(ext_rec.get(EXT_RAW_FIELDS[name][key], float("nan")), 4))
            row.append("")
        row += [""] * len(EXT_CHG_NAMES)
        row.append("")
        row += [""] * len(EXT_DERIVED)
        row.append("")
        ws.append(row)
        for pcol, rcol in zip(price_cols, ret_cols):
            if i < len(dates_desc) - 1:
                ws[f"{rcol}{r}"] = (
                    f"=IF(COUNT({pcol}{r}:{pcol}{r + 1})=2,"
                    f"ROUND(({pcol}{r}/{pcol}{r + 1}-1)*100,4),\"\")"
                )
            else:
                ws[f"{rcol}{r}"] = ""

        vc = get_column_letter(vix_idx["close"])
        tc = get_column_letter(tnx_idx["close"])
        vchg = get_column_letter(chg_idx["VIX_Chg%"])
        ws.cell(row=r, column=chg_idx["VIX回报"], value=(
            f"=IF(COUNT({vc}{r}:{vc}{r + 1})=2,ROUND(({vc}{r}/{vc}{r + 1}-1)*100,4),\"\")"
        ))
        ws.cell(row=r, column=chg_idx["VIX_Chg%"], value=(
            f"=IF(COUNT({vc}{r}:{vc}{r + 1})=2,ROUND(({vc}{r}/{vc}{r + 1}-1)*100,4),\"\")"
        ))
        ws.cell(row=r, column=chg_idx["TNX回报"], value=(
            f"=IF(COUNT({tc}{r}:{tc}{r + 1})=2,ROUND(({tc}{r}/{tc}{r + 1}-1)*100,4),\"\")"
        ))
        ws.cell(row=r, column=chg_idx["TNX_ChgBp"], value=(
            f"=IF(COUNT({tc}{r}:{tc}{r + 1})=2,({tc}{r}-{tc}{r + 1})*100,\"\")"
        ))
        if i + 5 < len(dates_desc):
            ws.cell(row=r, column=chg_idx["VIX_5dChg"], value=(
                f"=IF(COUNT({vc}{r}:{vc}{r + 5})=6,ROUND(({vc}{r}/{vc}{r + 5}-1)*100,4),\"\")"
            ))
        if i + 20 <= de:
            ws.cell(row=r, column=chg_idx["VIX_20dVol"], value=(
                f'=IF(COUNT({vchg}{r}:{vchg}{r + 19})=20,IFERROR(ROUND(STDEV({vchg}{r}:{vchg}{r + 19}),4),""),"")'
            ))
        for name, a, b in (
            ("CreditSpread", etf_return_col["HYG"], etf_return_col["TLT"]),
            ("JNKSpread", etf_return_col["JNK"], etf_return_col["TLT"]),
            ("USDGoldRatio", etf_return_col["UUP"], etf_return_col["GLD"]),
            ("SectRotation", etf_return_col["XLK"], etf_return_col["XLF"]),
            ("YldCurveProxy", etf_return_col["TLT"], etf_return_col["TIP"]),
        ):
            ws.cell(row=r, column=derived_idx[name], value=f'=IFERROR({a}{r}-{b}{r},"")')
        if i + 20 <= de:
            sp = etf_return_col["SPY"]
            tl = etf_return_col["TLT"]
            ws.cell(row=r, column=derived_idx["StkBonCorr"], value=(
                f'=IF(COUNT({sp}{r}:{sp}{r + 19})=20,IFERROR(CORREL({sp}{r}:{sp}{r + 19},{tl}{r}:{tl}{r + 19}),""),"")'
            ))
        ws.cell(row=r, column=derived_idx["VIX_TNX_Ratio"], value=f"={vc}{r}/{tc}{r}")
        for t in fund_order:
            fund_col = colmap[t]
            for c_idx, (cond, horizon) in zip(strat_cols[t], per_fund[t]):
                if i < horizon:
                    ws.cell(row=r, column=c_idx, value="")
                else:
                    cond_expr = sample.condition_expr(
                        cond, t, r, colmap, fund_col, strategy_thr_refs[c_idx]
                    )
                    target = target_formula(fund_col, r, horizon)
                    ws.cell(row=r, column=c_idx, value=f"=IF({cond_expr},{target},\"\")")
            cols = [get_column_letter(c) for c in strat_cols[t]]
            expr = f"{cols[-1]}{r}"
            for col in reversed(cols[:-1]):
                expr = f"IF({col}{r}<>\"\",{col}{r},{expr})"
            ws.cell(row=r, column=merged_idx[t], value=f"={expr}")

    pct_cols = list(fund_return_col.values()) + list(etf_return_col.values())
    ext_pct = [
        get_column_letter(chg_idx[name])
        for name in ("VIX回报", "VIX_Chg%", "VIX_5dChg", "VIX_20dVol", "TNX回报")
    ]
    ext_pct += [
        get_column_letter(derived_idx[name])
        for name in ("CreditSpread", "JNKSpread", "USDGoldRatio", "SectRotation", "YldCurveProxy")
    ]
    ext_bp = get_column_letter(chg_idx["TNX_ChgBp"])
    ext_num = [
        get_column_letter(derived_idx["StkBonCorr"]),
        get_column_letter(derived_idx["VIX_TNX_Ratio"]),
    ]
    for t in fund_order:
        pct_cols += [get_column_letter(c) for c in strat_cols[t]]
        pct_cols.append(get_column_letter(merged_idx[t]))
    for r in range(ds, de + 1):
        for col in pct_cols + ext_pct:
            ws[f"{col}{r}"].number_format = '0.0000"%"'
        ws[f"{ext_bp}{r}"].number_format = '0.00"bp"'
        for col in ext_num:
            ws[f"{col}{r}"].number_format = "0.0000"

    for c in range(2, n_cols + 1):
        if c in blank_cols:
            continue
        ws.cell(row=2, column=c).number_format = "0.00%"
        for r in (5, 6, 7):
            ws.cell(row=r, column=c).number_format = '0.0000"%"'
    for t in fund_order:
        for c_idx in strat_cols[t]:
            for r in (11, 12, 13):
                ws.cell(row=r, column=c_idx).number_format = 'General"%"'

    for r in range(2, 11):
        ws.cell(row=r, column=1).font = head_font
        for c in range(2, n_cols + 1):
            ws.cell(row=r, column=c).font = body_font
    for row in (14, 15):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = head_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[14].height = 47
    ws.row_dimensions[15].height = 47
    for r in range(ds, de + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = body_font
    for r in range(2, de + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=(r in (14, 15)),
            )
    ws.column_dimensions["A"].width = 12
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "B16"

    import os
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)
    print("saved:", out_path, "cols", n_cols, "rows", len(dates_desc))


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    master = s4.load_master()
    mapping = load_mapping()
    dates_all = sorted(pd.to_datetime(master["Date"]), reverse=True)
    dates_frozen = [d for d in dates_all if d >= pd.Timestamp(CUTOFF)]
    out_dir = config.RESULT_ROOT / "v4_正式版_m30_新版式"
    out_dir.mkdir(parents=True, exist_ok=True)
    build_workbook(out_dir / "全历史.xlsx", dates_all, mapping, master, "基金全历史")
    build_workbook(out_dir / "25-26年.xlsx", dates_frozen, mapping, master, "25-26年")


if __name__ == "__main__":
    main()
