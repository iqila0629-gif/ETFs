"""Patch the note sheet of existing m30 new-format workbooks in place."""

from __future__ import annotations

import pathlib
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font


NOTE_LINES = [
    "公式说明：",
    "1. 统计头公式（第2-10行）",
    "   Hit Ratio = Up Count / (Up Count + Down Count)；",
    "   Up Count = COUNTIF(该列数据区, \">0\")；Down Count = COUNTIF(该列数据区, \"<0\")；",
    "   Average = AVERAGE(该列数据区)；Max/Min/Count/Std/Sum 对应 MAX/MIN/COUNT/STDEV/SUM。",
    "2. 回报率公式",
    "   =IF(COUNT(B16:B17)=2,ROUND((B16/B17-1)*100,4),\"\")",
    "   B16 是今日 Adj Close，B17 是前一交易日 Adj Close；",
    "   COUNT 判断两天价格都存在才计算当日回报率，否则留空，避免 #DIV/0!。",
    "   VIX_Close 的回报率公式相同；VIX_Chg% 为外部涨跌幅列，可与计算值核对。",
    "   ETF 原始数据列为 Open/High/Low/Close/Adj Close/Volume，回报率由 Adj Close 计算；",
    "   第14行为区块名称行（如EEM），第15行为列头；ETF列头不重复ETF名。",
    "   表内按区块排列：基金区、每支ETF、ETF区、VIX区结束后均用空白列分隔。",
    "   所有回报率均以百分比显示并保留4位小数（公式 ROUND 到4位，显示格式 0.0000%）。",
    "3. 策略公式",
    "   =IF(条件, 该基金次日实际回报, \"\")",
    "   条件引用回报率列和阈值参数；满足条件时显示次日实际回报（日期降序，次日=上一行）。",
    "   horizon>1 时取未来 N 日平均回报，目标公式带 IFERROR，基金无历史数据时留空。",
    "4. 合并公式",
    "   策略列按 |全历史Average| 从高到低排列，合并列取第一个非空策略：",
    "   =IF(策略1<>\"\",策略1,IF(策略2<>\"\",策略2,策略3))",
    "   同一天多条策略触发时，历史 |Average| 最大的策略生效；全部未触发则留空。",
    "5. 阈值参数",
    "   每个策略列正上方的第11/12/13行是该策略条件阈值：二条件用12/13行，三条件额外用11行；",
    "   修改数字可调整条件，公式自动更新。",
]


def patch(path: pathlib.Path) -> None:
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    data = wb.worksheets[1]
    title = ws["A1"].value or "全量 m30 新版式"
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    for i, line in enumerate(NOTE_LINES, start=3):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 120
    ds = 16
    de = data.max_row
    for c in range(2, data.max_column + 1):
        header = data.cell(row=15, column=c).value
        if header in (None, ""):
            continue
        col = data.cell(row=15, column=c).column_letter
        data[f"{col}2"] = f"={col}3/({col}3+{col}4)"
        data[f"{col}5"] = f"=AVERAGE({col}{ds}:{col}{de})"
        data[f"{col}6"] = f"=MAX({col}{ds}:{col}{de})"
        data[f"{col}7"] = f"=MIN({col}{ds}:{col}{de})"
        data[f"{col}9"] = f"=STDEV({col}{ds}:{col}{de})"
        data[f"{col}10"] = f"=SUM({col}{ds}:{col}{de})"
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)
    print("patched:", path)


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if len(sys.argv) > 1:
        paths = [pathlib.Path(a) for a in sys.argv[1:]]
    else:
        root = pathlib.Path(__file__).resolve().parents[2] / "04_结果" / "v4_正式版_m30_新版式"
        paths = [
            root / "全历史.xlsx",
            root / "25-26年.xlsx",
        ]
    for p in paths:
        if p.exists():
            patch(p)


if __name__ == "__main__":
    main()
