"""Build the all-signal R4 merged reference version."""

from __future__ import annotations

import pathlib
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config
import scan_v4_thresholds as s4
from select_v4_strategies import signal_trades


CUTOFF = np.datetime64("2025-01-01")
STAT_NAMES = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")
title_font = Font(name="Arial", size=14, bold=True)
head_font = Font(name="Arial", size=10, bold=True)
body_font = Font(name="Arial", size=10)


def merge_daily(recs: list[dict]) -> tuple[dict, dict]:
    all_dates = set()
    for r in recs:
        all_dates |= r["dates"]
    by_date = {}
    for r in recs:
        for d, v in r["returns"].items():
            if d in all_dates and (d not in by_date or r["strength"] > by_date[d][1]):
                by_date[d] = (v, r["strength"])
    ordered = sorted(all_dates)
    tv = np.array([by_date[d][0] for d in ordered], dtype=float)
    td = np.array(ordered, dtype="datetime64[ns]")
    hold = td >= CUTOFF
    stats = {
        "full_avg": float(tv.mean()),
        "full_trades": int(tv.size),
        "full_hit": float((tv > 0).mean()),
        "frozen_avg": float(tv[hold].mean()) if hold.any() else float("nan"),
        "frozen_trades": int(hold.sum()),
        "frozen_hit": float((tv[hold] > 0).mean()) if hold.any() else float("nan"),
    }
    daily = {d: float(v) for d, (v, _s) in by_date.items()}
    return stats, daily


def write_company_wide(out_path: pathlib.Path, dates_desc: list[pd.Timestamp],
                       fund_order: list[str], fund_labels: dict[str, str],
                       daily: dict[str, dict], title: str) -> None:
    wb = Workbook()
    ws_note = wb.active
    ws_note.title = "说明"
    ws_note["A1"] = "v4 全信号 R4 合并参考版"
    ws_note["A1"].font = title_font
    for i, line in enumerate(
        [
            f"版本：{title}",
            "每支基金使用该基金全部正式信号，同日按 |全历史 Average| 最大者合并（R4）。",
            "注意：全信号合并会被弱信号稀释，只有部分基金整体 Average 仍 >= 0.2%，本表仅供对照。",
        ],
        start=3,
    ):
        ws_note.cell(row=i, column=1, value=line)
    ws_note.column_dimensions["A"].width = 110

    headers = ["日期"] + [fund_labels[t] for t in fund_order]
    ws = wb.create_sheet("数据")
    ws.append([])
    for i, label in enumerate(STAT_NAMES, start=2):
        ws.cell(row=i, column=1, value=label)
    ws.append([])
    ws.append([])
    ws.append(headers)
    ds = 14
    de = ds + len(dates_desc) - 1
    for col_idx in range(2, len(headers) + 1):
        col = get_column_letter(col_idx)
        ws[f"{col}2"] = f"={col}3/({col}3+{col}4)"
        ws[f"{col}3"] = f'=COUNTIF({col}{ds}:{col}{de},">0")'
        ws[f"{col}4"] = f'=COUNTIF({col}{ds}:{col}{de},"<0")'
        ws[f"{col}5"] = f"=AVERAGE({col}{ds}:{col}{de})"
        ws[f"{col}6"] = f"=MAX({col}{ds}:{col}{de})"
        ws[f"{col}7"] = f"=MIN({col}{ds}:{col}{de})"
        ws[f"{col}8"] = f"=COUNT({col}{ds}:{col}{de})"
        ws[f"{col}9"] = f"=STDEV({col}{ds}:{col}{de})"
        ws[f"{col}10"] = f"=SUM({col}{ds}:{col}{de})"
    for r in range(2, 11):
        ws.cell(row=r, column=1).font = head_font
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).font = body_font
    for r in range(13, 14):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = head_font
            ws.cell(row=r, column=c).fill = header_fill
    for d in dates_desc:
        row = [d.strftime("%m/%d/%Y")]
        for t in fund_order:
            row.append(round(daily[t][d], 4) if d in daily[t] else "")
        ws.append(row)
        r = ws.max_row
        ws.cell(row=r, column=1).border = border
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
    ws.column_dimensions["A"].width = 12
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24
    ws.freeze_panes = "B14"
    wb.save(out_path)
    print("saved:", out_path)


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    config.V4_OUT.mkdir(parents=True, exist_ok=True)
    delivery = config.RESULT_ROOT / "v4_正式交付_全信号"
    delivery.mkdir(parents=True, exist_ok=True)

    master = s4.load_master()
    fund_cols = list(pd.read_csv(config.FUND_PANEL).columns[1:])
    fund_set = set(fund_cols)
    external_cols = set(pd.read_csv(config.V4_EXTERNAL_PANEL).columns) - {"Date"}
    non_fund = {"Date"} | external_cols
    all_etfs = {c for c in master.columns if c not in fund_set and c not in non_fund}
    dates = master["Date"].to_numpy()
    dates_all = sorted(pd.to_datetime(dates), reverse=True)
    dates_frozen = [d for d in dates_all if d >= pd.Timestamp(CUTOFF)]

    pool = pd.read_csv(config.V4_FINAL20_COMBINED_PASS, keep_default_na=False)
    pool = pool.drop_duplicates(["ticker", "condition", "horizon"], keep="first")
    name_map = pd.read_csv(config.MIDDLE / "通用" / "文件" / "基金名称映射.csv", keep_default_na=False)
    name_by_ticker = dict(zip(name_map["ticker"], name_map["name"]))
    fund_order = sorted(pool["ticker"].unique())
    fund_labels = {t: f"{name_by_ticker.get(t, t)}（{t}）" for t in fund_order}

    target_cache: dict[tuple[str, int, bool], np.ndarray] = {}
    stats_rows = []
    daily = {}
    for ticker, grp in pool.groupby("ticker", sort=True):
        recs = []
        for r in grp.itertuples(index=False):
            rec = signal_trades(
                master, dates, ticker, str(r.condition), int(r.horizon),
                str(r.source), all_etfs, target_cache,
            )
            if rec is not None:
                recs.append(rec)
        st, d = merge_daily(recs)
        formal_pass = bool(
            abs(st["full_avg"]) >= 0.2
            and st["full_trades"] >= config.RECOMMENDED_FULL_TRADES
            and st["full_hit"] > 0.55
            and st["frozen_avg"] == st["frozen_avg"]
            and abs(st["frozen_avg"]) >= 0.2
            and st["frozen_trades"] >= config.RECOMMENDED_FROZEN_TRADES
            and st["frozen_hit"] >= 0.55
        )
        stats_rows.append({"ticker": ticker, "formal_pass": formal_pass, **st})
        daily[ticker] = d
    stats = pd.DataFrame(stats_rows).sort_values("ticker").reset_index(drop=True)
    stats.to_csv(config.V4_OUT / "v4_all_signal_merged_stats.csv", index=False)
    print("all-signal merged funds:", len(stats), "formal pass:", int(stats["formal_pass"].sum()))

    write_company_wide(delivery / "v4_公司格式_全信号_全历史.xlsx", dates_all, fund_order, fund_labels, daily, "全信号 R4 · 全历史")
    write_company_wide(delivery / "v4_公司格式_全信号_冻结期.xlsx", dates_frozen, fund_order, fund_labels, daily, "全信号 R4 · 冻结期 2025-2026")


if __name__ == "__main__":
    main()
