"""v4 Phase 5: build color-coded company delivery workbooks."""

from __future__ import annotations

import pathlib
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import scan_v4_thresholds as s4
import scan_v4_conditions as sc
from select_v4_strategies import signal_trades


CUTOFF = np.datetime64("2025-01-01")
STAT_NAMES = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
PALETTE = [
    "FFF2CC", "DDEBF7", "E2EFDA", "FCE4D6", "E4DFEC",
    "FFD966", "C6E0B4", "BDD7EE", "F8CBAD", "D9E1F2",
]

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")
title_font = Font(name="Arial", size=14, bold=True)
head_font = Font(name="Arial", size=10, bold=True)
body_font = Font(name="Arial", size=10)


def fund_color(ticker: str, idx: int) -> str:
    return PALETTE[idx % len(PALETTE)]


def write_note_sheet(ws, lines: list[str]) -> None:
    ws["A1"] = "正式交付说明"
    ws["A1"].font = title_font
    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110


def write_stats_header(ws, headers: list[str], max_strat: int, ds: int, de: int) -> None:
    ws.append([])
    for i, label in enumerate(STAT_NAMES, start=2):
        ws.cell(row=i, column=1, value=label)
    ws.append([])
    ws.append([])
    ws.append(headers)
    for _ in range(max_strat):
        ws.append([""] + [""] * (len(headers) - 1))
    for col_idx in range(2, len(headers) + 1):
        col = get_column_letter(col_idx)
        ws[f"{col}2"] = f"={col}3/({col}3+{col}4)"
        ws[f"{col}3"] = f'=COUNTIF({col}{ds}:{col}{de},">0")'
        ws[f"{col}4"] = f'=COUNTIF({col}{ds}:{col}{de},"<0")'
        ws[f"{col}5"] = f"=AVERAGE({col}{ds}:{col}{de})"
        ws[f"{col}6"] = f"=MAX({col}{ds}:{col}{de})"
        ws[f"{col}7"] = f"=MIN({col}{ds}:{col}{de})"
        ws[f"{col}8"] = f"=COUNT({col}{ds}:{col}{de})"
        ws[f"{col}9"] = f"=STDEV({col}{ds}:{col}{de})"
        ws[f"{col}10"] = f"=SUM({col}{ds}:{col}{de})"


def style_header_block(ws, n_cols: int, max_strat: int) -> None:
    for r in range(2, 11):
        ws.cell(row=r, column=1).font = head_font
        for c in range(2, n_cols + 1):
            ws.cell(row=r, column=c).font = body_font
    for r in range(13, 14 + max_strat):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = head_font if r == 13 else body_font
        if r == 13:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = header_fill


def build_company_workbook(
    out_path: pathlib.Path,
    dates_desc: list[pd.Timestamp],
    fund_order: list[str],
    fund_labels: dict[str, str],
    trade_data: dict[str, list[dict]],
    max_strat: int,
    etf_cols: list[str],
    master: pd.DataFrame,
    all_etfs: set[str],
    etf_related: dict[str, list[tuple[str, str]]],
    ext_columns: list[str],
    ext_related: dict[str, list[tuple[str, str]]],
    title: str,
) -> None:
    wb = Workbook()
    ws_note = wb.active
    ws_note.title = "说明"
    ds = 14 + max_strat
    de = ds + len(dates_desc) - 1
    write_note_sheet(
        ws_note,
        [
            f"版本：{title}",
            "【行列含义】基金数据 Sheet：",
            "1-12行：公司13行统计头（第1行空白；第2-10行=Hit Ratio/Up Count/Down Count/Average/Max/Min/Count/Std/Sum；第11-12行空白）。",
            "第13行：列名；A列=日期，其余列=基金名称（代码）。",
            f"第14-{13 + max_strat}行：策略色块行，每行一条策略并填色。",
            f"第{ds}行起：数据行，日期降序；基金列中日期格子按当天生效策略填色，格子内=触发日对应窗口的实际回报（%），空白=未触发。",
            "【行列含义】ETF外部示例 Sheet：",
            "第13行：A列=日期，其余列=20支ETF + 相关外部数据列（VIX/TNX/派生指标等）。",
            f"第14-{13 + max_strat}行：ETF/外部列相关策略色块行。",
            f"第{ds}行起：数据行；ETF/外部列中触发策略的日期格子填色。",
            "【策略命名规则】",
            "单条件：SPY_up / SPY_down / SPY_big_up / SPY_big_down / SPY_gt2 / SPY_lt-2 / SPY_bin_1_2。",
            "双条件：EEM_down_GDX_up 表示 EEM 跌且 GDX 涨。",
            "三条件：FXY_up_GDX_up_QQQ_down 表示 FXY 涨、GDX 涨、QQQ 跌，全部为且的关系。",
            "复合条件：combo_QQQ_down__self_big_down 表示 QQQ 跌且基金自身大跌。",
            "外部条件：ext_vix_chg_ge5 表示 VIX 当日变化 >=5%；ext_vix_close_ge25 表示 VIX 收盘 >=25。",
            "自身条件：self_3down 表示基金自身连续3日下跌。",
            "预测窗口：horizon=N 表示未来 N 个交易日，多日回报取未来 N 日日平均。",
            "同日多策略触发：按 |全历史 Average| 最大者着色；颜色为基金/ETF 内部调色。",
        ],
    )
    headers = ["日期"] + [fund_labels[t] for t in fund_order]
    ws = wb.create_sheet("基金数据")
    write_stats_header(ws, headers, max_strat, ds, de)
    for i, t in enumerate(fund_order):
        col_idx = i + 2
        recs = trade_data[t]
        for j, rec in enumerate(recs):
            ws.cell(row=14 + j, column=col_idx, value=rec["condition"]).fill = PatternFill(
                "solid", fgColor=fund_color(t, j)
            )
    for d in dates_desc:
        row = [d.strftime("%m/%d/%Y")]
        for t in fund_order:
            best = None
            for j, rec in enumerate(trade_data[t]):
                if d in rec["dates"] and (best is None or rec["strength"] > best[0]["strength"]):
                    best = (rec, j)
            if best is None:
                row.append("")
            else:
                rec, j = best
                row.append(round(float(rec["returns"][d]), 4))
        ws.append(row)
        r = ws.max_row
        ws.cell(row=r, column=1).border = border
        for t_idx, t in enumerate(fund_order):
            c = t_idx + 2
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if cell.value not in (None, ""):
                best = None
                for j, rec in enumerate(trade_data[t]):
                    if d in rec["dates"] and (best is None or rec["strength"] > best[0]["strength"]):
                        best = (rec, j)
                if best:
                    cell.fill = PatternFill("solid", fgColor=fund_color(t, best[1]))
    style_header_block(ws, len(headers), max_strat)
    ws.column_dimensions["A"].width = 12
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24
    ws.freeze_panes = f"B{ds}"

    etf_headers = ["日期"] + etf_cols + ext_columns
    ws2 = wb.create_sheet("ETF外部示例")
    write_stats_header(ws2, etf_headers, max_strat, ds, de)
    for c_idx, code in enumerate(etf_cols, start=2):
        rel = etf_related.get(code, [])
        for j, (_, cond) in enumerate(rel):
            ws2.cell(row=14 + j, column=c_idx, value=cond).fill = PatternFill(
                "solid", fgColor=fund_color(code, j)
            )
    for c_idx, col in enumerate(ext_columns, start=2 + len(etf_cols)):
        rel = ext_related.get(col, [])
        for j, (_, cond) in enumerate(rel):
            ws2.cell(row=14 + j, column=c_idx, value=cond).fill = PatternFill(
                "solid", fgColor=fund_color(col, j)
            )
    ext_map = {
        col: dict(zip(pd.to_datetime(master["Date"]), master[col]))
        for col in ext_columns
    }
    etf_value_map = {
        code: dict(zip(pd.to_datetime(master["Date"]), master[code])) for code in etf_cols
    }
    for d in dates_desc:
        row = [d.strftime("%m/%d/%Y")]
        for code in etf_cols:
            val = etf_value_map[code].get(d)
            row.append("" if val is None or pd.isna(val) else round(float(val), 4))
        for col in ext_columns:
            val = ext_map[col].get(d)
            row.append("" if val is None or pd.isna(val) else round(float(val), 4))
        ws2.append(row)
        r = ws2.max_row
        ws2.cell(row=r, column=1).border = border
        for c_idx, code in enumerate(etf_cols, start=2):
            cell = ws2.cell(row=r, column=c_idx)
            cell.border = border
            if cell.value in (None, ""):
                continue
            best = None
            for j, (t, cond) in enumerate(etf_related.get(code, [])):
                rec = next((x for x in trade_data[t] if x["condition"] == cond), None)
                if rec is not None and d in rec["dates"] and (best is None or rec["strength"] > best[0]["strength"]):
                    best = (rec, j)
            if best:
                cell.fill = PatternFill("solid", fgColor=fund_color(code, best[1]))
        for c_idx, col in enumerate(ext_columns, start=2 + len(etf_cols)):
            cell = ws2.cell(row=r, column=c_idx)
            cell.border = border
            if cell.value in (None, ""):
                continue
            best = None
            for j, (t, cond) in enumerate(ext_related.get(col, [])):
                rec = next((x for x in trade_data[t] if x["condition"] == cond), None)
                if rec is not None and d in rec["dates"] and (best is None or rec["strength"] > best[0]["strength"]):
                    best = (rec, j)
            if best:
                cell.fill = PatternFill("solid", fgColor=fund_color(col, best[1]))
        for c_idx in range(2 + len(etf_cols), len(etf_headers) + 1):
            ws2.cell(row=r, column=c_idx).border = border
    style_header_block(ws2, len(etf_headers), max_strat)
    ws2.column_dimensions["A"].width = 12
    for c in range(2, len(etf_headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 12
    ws2.freeze_panes = "B14"

    wb.save(out_path)
    print("saved:", out_path)


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    config.V4_OUT.mkdir(parents=True, exist_ok=True)
    suffix = sys.argv[1] if len(sys.argv) > 1 else ""
    outdir = None
    if "--outdir" in sys.argv:
        idx = sys.argv.index("--outdir")
        outdir = sys.argv[idx + 1]
    delivery = config.RESULT_ROOT / outdir if outdir else config.RESULT_ROOT / f"v4_正式交付{suffix}"
    delivery.mkdir(parents=True, exist_ok=True)

    master = s4.load_master()
    fund_cols = list(pd.read_csv(config.FUND_PANEL).columns[1:])
    fund_set = set(fund_cols)
    external_cols = set(pd.read_csv(config.V4_EXTERNAL_PANEL).columns) - {"Date"}
    non_fund = {"Date"} | external_cols
    all_etfs = {c for c in master.columns if c not in fund_set and c not in non_fund}
    dates = master["Date"].to_numpy()
    dates_all = sorted(pd.to_datetime(dates), reverse=True)
    dates_frozen = [d for d in dates_all if d >= pd.Timestamp(CUTOFF)]

    mapping = pd.read_csv(config.V4_OUT / f"v4_strategy_mapping{suffix}.csv", keep_default_na=False)
    name_map = pd.read_csv(config.MIDDLE / "通用" / "文件" / "基金名称映射.csv", keep_default_na=False)
    name_by_ticker = dict(zip(name_map["ticker"], name_map["name"]))
    fund_order = sorted(mapping["ticker"].unique())
    fund_labels = {t: f"{name_by_ticker.get(t, t)}（{t}）" for t in fund_order}

    target_cache: dict[tuple[str, int, bool], np.ndarray] = {}
    trade_data: dict[str, list[dict]] = {}
    etf_related: dict[str, list[tuple[str, str]]] = {c: [] for c in all_etfs}
    ext_related: dict[str, list[tuple[str, str]]] = {}

    def ext_part_to_col(part: str) -> str | None:
        for safe, col in sc.EXTERNAL_COLS.items():
            if part.startswith(f"ext_{safe}_"):
                return col
        return None

    for ticker, grp in mapping.groupby("ticker", sort=True):
        recs = []
        for r in grp.sort_values("strategy_no").itertuples(index=False):
            rec = signal_trades(
                master, dates, ticker, str(r.condition), int(r.horizon),
                str(r.source), all_etfs, target_cache,
            )
            if rec is not None:
                recs.append(rec)
                tokens = {t for t in str(r.condition).split("_") if t in all_etfs}
                for code in tokens:
                    etf_related[code].append((ticker, str(r.condition)))
                parts = (
                    str(r.condition).split("__")
                    if "__" in str(r.condition)
                    else [str(r.condition)]
                )
                for part in parts:
                    col = ext_part_to_col(part)
                    if col is not None:
                        ext_related.setdefault(col, []).append((ticker, str(r.condition)))
        trade_data[ticker] = recs
    max_strat = mapping.groupby("ticker").size().max()
    etf_cols = sorted(all_etfs)
    ext_order = [
        "VIX_Close", "VIX_Chg%", "VIX_5dChg", "VIX_20dVol",
        "TNX_Yield", "TNX_ChgBp", "CreditSpread", "JNKSpread",
        "StkBonCorr", "USDGoldRatio", "SectRotation", "VIX_TNX_Ratio",
        "YldCurveProxy",
    ]
    ext_columns = [c for c in ext_order if c in ext_related or c in ("VIX_Chg%", "TNX_ChgBp")]
    print("funds", len(fund_order), "max strategies", max_strat, "etfs", len(etf_cols), flush=True)

    build_company_workbook(
        delivery / "v4_公司格式_最佳策略_全历史.xlsx",
        dates_all, fund_order, fund_labels, trade_data, max_strat,
        etf_cols, master, all_etfs, etf_related, ext_columns, ext_related, "最佳策略 · 全历史",
    )
    build_company_workbook(
        delivery / "v4_公司格式_最佳策略_冻结期.xlsx",
        dates_frozen, fund_order, fund_labels, trade_data, max_strat,
        etf_cols, master, all_etfs, etf_related, ext_columns, ext_related, "最佳策略 · 冻结期 2025-2026",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_note_sheet(
        ws,
        [
            "每支基金最多 3-5 条策略；策略合并口径为同日取 |全历史 Average| 最大者。",
            "【行列含义】数据 Sheet：每行一条策略。",
            "列含义：基金名称（代码）、策略序号、触发条件、预测窗口（日）、全历史平均回报（%）、全历史交易数、全历史命中率、冻结期平均回报（%）、冻结期交易数、冻结期命中率。",
            "策略命名规则：单条件如 SPY_up；双条件如 EEM_down_GDX_up；三条件如 FXY_up_GDX_up_QQQ_down；复合条件如 combo_QQQ_down__self_big_down；外部条件如 ext_vix_chg_ge5；自身条件如 self_3down。",
        ],
    )
    ws2 = wb.create_sheet("数据")
    headers = ["基金名称（代码）", "策略序号", "触发条件", "预测窗口（日）", "全历史平均回报（%）",
               "全历史交易数", "全历史命中率", "冻结期平均回报（%）", "冻结期交易数", "冻结期命中率"]
    ws2.append(headers)
    for r in mapping.sort_values(["ticker", "strategy_no"]).itertuples(index=False):
        ws2.append(
            [
                fund_labels.get(r.ticker, r.ticker),
                r.strategy_no,
                r.condition,
                r.horizon,
                r.full_avg,
                r.full_trades,
                r.full_hit,
                r.frozen_avg,
                r.frozen_trades,
                r.frozen_hit,
            ]
        )
    for cell in ws2[1]:
        cell.font = head_font
    widths = [34, 10, 36, 12, 16, 14, 14, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    wb.save(delivery / "v4_每基金策略映射.xlsx")
    print("saved:", delivery / "v4_每基金策略映射.xlsx")

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "说明"
    write_note_sheet(
        ws3,
        [
            "最佳策略逐日明细：日期为触发日，实际回报为该策略对应窗口回报。",
            "【行列含义】数据 Sheet：每行一次触发。",
            "列含义：基金名称（代码）、触发条件、预测窗口（日）、日期、实际回报（%）。",
            "日期格式：MM/DD/YYYY，降序。",
        ],
    )
    ws4 = wb3.create_sheet("数据")
    headers4 = ["基金名称（代码）", "触发条件", "预测窗口（日）", "日期", "实际回报（%）"]
    ws4.append(headers4)
    for ticker in fund_order:
        label = fund_labels[ticker]
        for rec in trade_data[ticker]:
            for d, v in sorted(rec["returns"].items(), reverse=True):
                ws4.append([label, rec["condition"], rec["horizon"], d.strftime("%m/%d/%Y"), round(float(v), 4)])
    for cell in ws4[1]:
        cell.font = head_font
    for i, w in enumerate([34, 36, 12, 12, 14], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    wb3.save(delivery / "v4_信号逐日明细_最佳策略.xlsx")
    print("saved:", delivery / "v4_信号逐日明细_最佳策略.xlsx")


if __name__ == "__main__":
    main()
