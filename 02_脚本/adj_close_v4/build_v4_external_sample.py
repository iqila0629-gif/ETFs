"""Build a small sample showing the new external-data layout."""

from __future__ import annotations

import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import build_v4_m30_newformat_sample as sample


ETF_ALL = ["SPY", "TLT", "HYG", "JNK", "UUP", "GLD", "XLK", "XLF", "TIP"]
DATES = [
    "2026-08-07", "2026-08-06", "2026-08-05",
    "2026-08-04", "2026-08-03", "2026-07-31", "2026-07-30", "2026-07-29",
    "2026-07-28", "2026-07-27", "2026-07-24", "2026-07-23", "2026-07-22",
    "2026-07-21", "2026-07-20", "2026-07-17", "2026-07-16", "2026-07-15",
    "2026-07-14", "2026-07-13", "2026-07-10", "2026-07-09", "2026-07-08",
    "2026-07-07", "2026-07-06",
]

VIX = {
    "2026-08-11": (15.42, 15.61, 15.23, 15.28, 15.28, 0),
    "2026-08-10": (15.40, 15.72, 15.10, 15.46, 15.46, 0),
    "2026-08-07": (15.30, 15.36, 14.77, 14.90, 14.90, 0),
    "2026-08-06": (15.83, 16.03, 15.11, 15.15, 15.15, 0),
    "2026-08-05": (16.15, 18.43, 15.48, 15.81, 15.81, 0),
    "2026-08-04": (15.76, 16.65, 15.51, 16.50, 16.50, 0),
    "2026-08-03": (16.03, 16.30, 15.54, 15.86, 15.86, 0),
    "2026-07-31": (16.82, 18.70, 15.82, 15.99, 15.99, 0),
    "2026-07-30": (19.56, 20.08, 17.00, 17.09, 17.09, 0),
    "2026-07-29": (18.27, 20.88, 17.45, 20.66, 20.66, 0),
    "2026-07-28": (19.05, 19.52, 17.88, 18.21, 18.21, 0),
    "2026-07-27": (17.62, 19.93, 17.53, 18.67, 18.67, 0),
    "2026-07-24": (18.96, 19.05, 17.41, 18.58, 18.58, 0),
    "2026-07-23": (17.67, 20.31, 17.32, 18.70, 18.70, 0),
    "2026-07-22": (17.42, 19.49, 16.64, 16.64, 16.64, 0),
    "2026-07-21": (17.48, 17.99, 16.86, 17.05, 17.05, 0),
    "2026-07-20": (18.90, 18.94, 17.41, 18.65, 18.65, 0),
    "2026-07-17": (18.01, 19.50, 17.68, 18.77, 18.77, 0),
    "2026-07-16": (15.82, 17.23, 15.77, 16.73, 16.73, 0),
    "2026-07-15": (16.20, 16.57, 15.64, 15.67, 15.67, 0),
    "2026-07-14": (17.21, 17.56, 16.15, 16.50, 16.50, 0),
    "2026-07-13": (16.32, 17.41, 16.03, 17.16, 17.16, 0),
    "2026-07-10": (16.06, 16.16, 14.96, 15.03, 15.03, 0),
    "2026-07-09": (16.58, 17.27, 15.76, 15.84, 15.84, 0),
    "2026-07-08": (16.55, 18.91, 16.35, 16.90, 16.90, 0),
    "2026-07-07": (15.87, 16.64, 15.53, 16.13, 16.13, 0),
    "2026-07-06": (16.40, 16.50, 15.56, 15.57, 15.57, 0),
}

TNX = {
    "2026-08-11": (4.692, 4.703, 4.668, 4.684, 4.684, 0),
    "2026-08-10": (4.662, 4.703, 4.662, 4.699, 4.699, 0),
    "2026-08-07": (4.664, 4.672, 4.603, 4.660, 4.660, 0),
    "2026-08-06": (4.643, 4.676, 4.635, 4.670, 4.670, 0),
    "2026-08-05": (4.609, 4.639, 4.607, 4.617, 4.617, 0),
    "2026-08-04": (4.665, 4.668, 4.619, 4.627, 4.627, 0),
    "2026-08-03": (4.671, 4.702, 4.671, 4.686, 4.686, 0),
    "2026-07-31": (4.684, 4.747, 4.682, 4.745, 4.745, 0),
    "2026-07-30": (4.669, 4.686, 4.651, 4.663, 4.663, 0),
    "2026-07-29": (4.637, 4.655, 4.610, 4.622, 4.622, 0),
    "2026-07-28": (4.630, 4.635, 4.588, 4.604, 4.604, 0),
    "2026-07-27": (4.647, 4.665, 4.630, 4.641, 4.641, 0),
    "2026-07-24": (4.683, 4.691, 4.651, 4.679, 4.679, 0),
    "2026-07-23": (4.699, 4.714, 4.693, 4.703, 4.703, 0),
    "2026-07-22": (4.640, 4.661, 4.632, 4.657, 4.657, 0),
    "2026-07-21": (4.600, 4.640, 4.598, 4.628, 4.628, 0),
    "2026-07-20": (4.566, 4.608, 4.560, 4.598, 4.598, 0),
    "2026-07-17": (4.521, 4.550, 4.511, 4.541, 4.541, 0),
    "2026-07-16": (4.582, 4.596, 4.561, 4.569, 4.569, 0),
    "2026-07-15": (4.608, 4.610, 4.539, 4.545, 4.545, 0),
    "2026-07-14": (4.614, 4.614, 4.525, 4.585, 4.585, 0),
    "2026-07-13": (4.585, 4.618, 4.573, 4.609, 4.609, 0),
    "2026-07-10": (4.541, 4.571, 4.539, 4.569, 4.569, 0),
    "2026-07-09": (4.577, 4.581, 4.529, 4.539, 4.539, 0),
    "2026-07-08": (4.561, 4.597, 4.557, 4.569, 4.569, 0),
    "2026-07-07": (4.497, 4.533, 4.485, 4.529, 4.529, 0),
    "2026-07-06": (4.457, 4.491, 4.457, 4.479, 4.479, 0),
}

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")
head_font = Font(name="Arial", size=9, bold=True)
body_font = Font(name="Arial", size=9)


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    wb = Workbook()
    ws_note = wb.active
    ws_note.title = "说明"
    ws_note["A1"] = "外部数据新版式示例"
    ws_note["A1"].font = Font(name="Arial", size=14, bold=True)
    note_lines = [
        "布局说明：",
        "1. VIX/TNX 原始数据按 ETF 格式展示：第14行=名称，第15行=Open/High/Low/Close/Adj Close/Volume。",
        "2. VIX/TNX 回报与变化指标集中放一起，第15行直接写名称（如 VIX回报、VIX_Chg%、TNX_ChgBp），后面用空列分隔。",
        "3. 其他派生外部列：第14行=指标名称，第15行=计算公式（谁减谁），数据区为 Excel 公式。",
        "4. ETF 回报、VIX/TNX 回报、派生指标均用公式计算，回报率保留4位小数。",
    ]
    for i, line in enumerate(note_lines, start=3):
        ws_note.cell(row=i, column=1, value=line)
    ws_note.column_dimensions["A"].width = 120

    ws = wb.create_sheet("数据")
    col_idx = 2
    blank_cols: set[int] = set()

    etf_cols = {}
    etf_ohlc = {e: sample.load_etf_ohlc(e) for e in ETF_ALL}
    for e in ETF_ALL:
        etf_cols[e] = {}
        for key in ("open", "high", "low", "close", "adj", "volume", "ret"):
            etf_cols[e][key] = col_idx
            col_idx += 1
        blank_cols.add(col_idx); col_idx += 1

    vix_cols = {}
    tnx_cols = {}
    for cols in (vix_cols, tnx_cols):
        for key in ("open", "high", "low", "close", "adj", "volume"):
            cols[key] = col_idx
            col_idx += 1
        blank_cols.add(col_idx); col_idx += 1

    chg_cols = {}
    for key in ("VIX回报", "VIX_Chg%", "VIX_5dChg", "VIX_20dVol", "TNX回报", "TNX_ChgBp"):
        chg_cols[key] = col_idx
        col_idx += 1
    blank_cols.add(col_idx); col_idx += 1

    derived = {
        "CreditSpread": "HYG-TLT",
        "JNKSpread": "JNK-TLT",
        "StkBonCorr": "CORREL(SPY,TLT,近20日)",
        "USDGoldRatio": "UUP-GLD",
        "SectRotation": "XLK-XLF",
        "VIX_TNX_Ratio": "VIX_close/TNX_close",
        "YldCurveProxy": "TLT-TIP",
    }
    derived_cols = {}
    for name in derived:
        derived_cols[name] = col_idx
        col_idx += 1
    n_cols = col_idx - 1

    ws.cell(row=14, column=1, value="日期")
    ws.cell(row=15, column=1, value="日期")
    for e in ETF_ALL:
        label_c = etf_cols[e]["open"] + 2
        ws.cell(row=14, column=label_c, value=e)
        for key, label in (
            ("open", "Open"), ("high", "High"), ("low", "Low"),
            ("close", "Close"), ("adj", "Adj Close"), ("volume", "Volume"),
            ("ret", "回报"),
        ):
            ws.cell(row=15, column=etf_cols[e][key], value=label)
    for name, cols in (("VIX", vix_cols), ("TNX", tnx_cols)):
        ws.cell(row=14, column=cols["open"] + 2, value=name)
        for key, label in (
            ("open", "Open"), ("high", "High"), ("low", "Low"),
            ("close", "Close"), ("adj", "Adj Close"), ("volume", "Volume"),
        ):
            ws.cell(row=15, column=cols[key], value=label)
    chg_formula = {
        "VIX_Chg%": "VIX_close/昨VIX_close-1",
        "VIX_5dChg": "VIX_close/5日前VIX_close-1",
        "VIX_20dVol": "STDEV(VIX日变化,近20日)",
        "TNX_ChgBp": "(TNX_close-昨TNX_close)*100",
    }
    for key in chg_cols:
        if key in chg_formula:
            ws.cell(row=14, column=chg_cols[key], value=key)
            ws.cell(row=15, column=chg_cols[key], value=chg_formula[key])
        else:
            ws.cell(row=15, column=chg_cols[key], value=key)
    for name, formula in derived.items():
        ws.cell(row=14, column=derived_cols[name], value=name)
        ws.cell(row=15, column=derived_cols[name], value=formula)

    ds = 16
    de = ds + len(DATES) - 1
    for i, d in enumerate(DATES):
        r = ds + i
        date = pd.Timestamp(d)
        ws.cell(row=r, column=1, value=d.replace("-", "/"))
        for e in ETF_ALL:
            rec = etf_ohlc[e].get(date, {})
            for key in ("open", "high", "low", "close", "adj", "volume"):
                v = rec.get(key)
                if v is not None:
                    ws.cell(row=r, column=etf_cols[e][key], value=v if key == "volume" else round(v, 4))
            adj = get_column_letter(etf_cols[e]["adj"])
            if i < len(DATES) - 1:
                ws.cell(row=r, column=etf_cols[e]["ret"], value=(
                    f"=IF(COUNT({adj}{r}:{adj}{r + 1})=2,"
                    f"ROUND(({adj}{r}/{adj}{r + 1}-1)*100,4),\"\")"
                ))
        for cols, data in ((vix_cols, VIX), (tnx_cols, TNX)):
            vals = data.get(d)
            if vals:
                for key, v in zip(("open", "high", "low", "close", "adj", "volume"), vals):
                    ws.cell(row=r, column=cols[key], value=v if key == "volume" else round(v, 4))

        vc = get_column_letter(vix_cols["close"])
        tc = get_column_letter(tnx_cols["close"])
        ws.cell(row=r, column=chg_cols["VIX回报"], value=(
            f"=IF(COUNT({vc}{r}:{vc}{r + 1})=2,ROUND(({vc}{r}/{vc}{r + 1}-1)*100,4),\"\")"
        ))
        ws.cell(row=r, column=chg_cols["VIX_Chg%"], value=(
            f"=IF(COUNT({vc}{r}:{vc}{r + 1})=2,ROUND(({vc}{r}/{vc}{r + 1}-1)*100,4),\"\")"
        ))
        ws.cell(row=r, column=chg_cols["TNX回报"], value=(
            f"=IF(COUNT({tc}{r}:{tc}{r + 1})=2,ROUND(({tc}{r}/{tc}{r + 1}-1)*100,4),\"\")"
        ))
        ws.cell(row=r, column=chg_cols["TNX_ChgBp"], value=(
            f"=IF(COUNT({tc}{r}:{tc}{r + 1})=2,({tc}{r}-{tc}{r + 1})*100,\"\")"
        ))
        if i + 5 < len(DATES):
            ws.cell(row=r, column=chg_cols["VIX_5dChg"], value=(
                f"=IF(COUNT({vc}{r}:{vc}{r + 5})=6,ROUND(({vc}{r}/{vc}{r + 5}-1)*100,4),\"\")"
            ))
        if i + 20 <= de:
            vchg = get_column_letter(chg_cols["VIX_Chg%"])
            ws.cell(row=r, column=chg_cols["VIX_20dVol"], value=(
                f'=IF(COUNT({vchg}{r}:{vchg}{r + 19})=20,IFERROR(ROUND(STDEV({vchg}{r}:{vchg}{r + 19}),4),""),"")'
            ))
        for name, a, b in (
            ("CreditSpread", etf_cols["HYG"]["ret"], etf_cols["TLT"]["ret"]),
            ("JNKSpread", etf_cols["JNK"]["ret"], etf_cols["TLT"]["ret"]),
            ("USDGoldRatio", etf_cols["UUP"]["ret"], etf_cols["GLD"]["ret"]),
            ("SectRotation", etf_cols["XLK"]["ret"], etf_cols["XLF"]["ret"]),
            ("YldCurveProxy", etf_cols["TLT"]["ret"], etf_cols["TIP"]["ret"]),
        ):
            a_letter = get_column_letter(a)
            b_letter = get_column_letter(b)
            ws.cell(row=r, column=derived_cols[name], value=f'=IFERROR({a_letter}{r}-{b_letter}{r},"")')
        if i + 20 <= de:
            sp = get_column_letter(etf_cols["SPY"]["ret"])
            tl = get_column_letter(etf_cols["TLT"]["ret"])
            ws.cell(row=r, column=derived_cols["StkBonCorr"], value=(
                f'=IF(COUNT({sp}{r}:{sp}{r + 19})=20,IFERROR(CORREL({sp}{r}:{sp}{r + 19},{tl}{r}:{tl}{r + 19}),""),"")'
            ))
        ws.cell(row=r, column=derived_cols["VIX_TNX_Ratio"], value=f"={vc}{r}/{tc}{r}")

    pct_cols = []
    for e in ETF_ALL:
        pct_cols.append(etf_cols[e]["ret"])
    pct_cols += [
        chg_cols["VIX回报"], chg_cols["VIX_Chg%"], chg_cols["VIX_5dChg"],
        chg_cols["VIX_20dVol"], chg_cols["TNX回报"],
    ]
    pct_cols += [
        derived_cols["CreditSpread"], derived_cols["JNKSpread"],
        derived_cols["USDGoldRatio"], derived_cols["SectRotation"],
        derived_cols["YldCurveProxy"],
    ]
    for r in range(ds, de + 1):
        for c in pct_cols:
            ws.cell(row=r, column=c).number_format = '0.0000"%"'
        ws.cell(row=r, column=chg_cols["TNX_ChgBp"]).number_format = '0.00"bp"'
        ws.cell(row=r, column=derived_cols["StkBonCorr"]).number_format = "0.0000"
        ws.cell(row=r, column=derived_cols["VIX_TNX_Ratio"]).number_format = "0.0000"

    for row in (14, 15):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = head_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 30
    for r in range(ds, de + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = body_font
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 12
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B16"
    wb.calculation.fullCalcOnLoad = True

    out_dir = config.RESULT_ROOT / "示例审批"
    out_path = out_dir / "外部数据新版式示例.xlsx"
    wb.save(out_path)
    print("saved:", out_path, "cols", n_cols, "rows", len(DATES))


if __name__ == "__main__":
    main()
