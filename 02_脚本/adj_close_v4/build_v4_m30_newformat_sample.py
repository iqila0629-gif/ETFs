"""Build a two-fund m30 sample in the auditable Excel-formula layout."""

from __future__ import annotations

import json
import os
import pathlib
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import scan_v4_thresholds as s4


STAT_NAMES = ["Hit Ratio", "Up Count", "Down Count", "Average", "Max", "Min", "Count", "Std", "Sum"]
FUNDS = ["UOPIX", "ULPIX"]
ETF_ALL = ["EEM", "GLD", "HYG", "TIP", "XLF", "FXY", "GDX", "XLC"]
EXT_RAW_COLS = ["VIX_Close"]
EXT_COLS_USED = ["VIX_Chg%"]
WINDOW_DAYS = 60

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")
title_font = Font(name="Arial", size=14, bold=True)
head_font = Font(name="Arial", size=9, bold=True)
body_font = Font(name="Arial", size=9)

def etf_cond_expr(tokens: list[str], row: int, colmap: dict[str, str], thr_ref: str) -> str:
    etf = tokens[0]
    suffix = "_".join(tokens[1:])
    col = colmap[etf]
    if suffix == "up":
        return f"{col}{row}>{thr_ref}"
    if suffix == "down":
        return f"{col}{row}<{thr_ref}"
    if suffix == "big_up":
        return f"{col}{row}>={thr_ref}"
    if suffix == "big_down":
        return f"{col}{row}<={thr_ref}"
    if suffix == "gt2":
        return f"{col}{row}>{thr_ref}"
    if suffix == "lt-2":
        return f"{col}{row}<{thr_ref}"
    if suffix.startswith("bin_"):
        band = suffix[4:]
        if band == "gt2":
            return f"{col}{row}>{thr_ref}"
        if band == "lt-2":
            return f"{col}{row}<{thr_ref}"
        lo, hi = (float(x) for x in band.split("_"))
        return f"AND({col}{row}>{lo},{col}{row}<={hi})"
    raise ValueError(suffix)


def ext_cond_expr(part: str, row: int, colmap: dict[str, str], thr_ref: str) -> str:
    import scan_v4_conditions as sc
    for safe, col_name in sc.EXTERNAL_COLS.items():
        prefix = f"ext_{safe}_"
        if part.startswith(prefix):
            op = part[len(prefix):]
            col = colmap[col_name]
            if op == "up":
                return f"{col}{row}>{thr_ref}"
            if op == "down":
                return f"{col}{row}<{thr_ref}"
            if op.startswith("ge"):
                return f"{col}{row}>={thr_ref}"
            return f"{col}{row}<={thr_ref}"
    raise ValueError(part)


def self_cond_expr(suffix: str, fund_col: str, row: int, thr_ref: str) -> str:
    if suffix == "up":
        return f"{fund_col}{row}>{thr_ref}"
    if suffix == "down":
        return f"{fund_col}{row}<{thr_ref}"
    if suffix == "big_up":
        return f"{fund_col}{row}>={thr_ref}"
    if suffix == "big_down":
        return f"{fund_col}{row}<={thr_ref}"
    if suffix in ("3up", "3down", "5up", "5down"):
        n = int(suffix[0])
        sign = ">" if suffix.endswith("up") else "<"
        parts = [f"{fund_col}{row - k}{sign}{thr_ref}" for k in range(n)]
        return "AND(" + ",".join(parts) + ")"
    raise ValueError(suffix)


def part_expr(part: str, ticker: str, row: int, colmap: dict[str, str], fund_col: str, thr_ref: str) -> str:
    if part.startswith("ext_"):
        return ext_cond_expr(part, row, colmap, thr_ref)
    if part.startswith("self_"):
        return self_cond_expr(part.removeprefix("self_"), fund_col, row, thr_ref)
    return etf_cond_expr(part.split("_"), row, colmap, thr_ref)


def condition_parts(condition: str, colmap: dict[str, str]) -> list[str]:
    if condition.startswith("combo_"):
        return condition[len("combo_"):].split("__")
    if condition.startswith("ext_") or condition.startswith("self_"):
        return [condition]
    tokens = condition.split("_")
    if len(tokens) == 6 and tokens[0] in colmap and tokens[2] in colmap and tokens[4] in colmap:
        return ["_".join(tokens[0:2]), "_".join(tokens[2:4]), "_".join(tokens[4:6])]
    if len(tokens) == 4 and tokens[0] in colmap and tokens[2] in colmap:
        return ["_".join(tokens[0:2]), "_".join(tokens[2:4])]
    return [condition]


def part_threshold(part: str) -> float:
    if part.startswith("ext_"):
        import scan_v4_conditions as sc
        for safe, col_name in sc.EXTERNAL_COLS.items():
            prefix = f"ext_{safe}_"
            if part.startswith(prefix):
                op = part[len(prefix):]
                if op in ("up", "down"):
                    return 0.0
                return float(op[2:].replace("_", "."))
        raise ValueError(part)
    if part.startswith("self_"):
        suffix = part.removeprefix("self_")
        if suffix in ("up", "down", "3up", "3down", "5up", "5down"):
            return 0.0
        if suffix == "big_up":
            return 2.0
        if suffix == "big_down":
            return -2.0
        raise ValueError(suffix)
    tokens = part.split("_")
    suffix = "_".join(tokens[1:])
    if suffix in ("up", "down"):
        return 0.0
    if suffix == "big_up":
        return 1.0
    if suffix == "big_down":
        return -1.0
    if suffix == "gt2":
        return 2.0
    if suffix == "lt-2":
        return -2.0
    raise ValueError(suffix)


def condition_expr(
    condition: str,
    ticker: str,
    row: int,
    colmap: dict[str, str],
    fund_col: str,
    thr_refs: list[str],
) -> str:
    parts = condition_parts(condition, colmap)
    exprs = [
        part_expr(p, ticker, row, colmap, fund_col, ref)
        for p, ref in zip(parts, thr_refs)
    ]
    if len(exprs) == 1:
        return exprs[0]
    return "AND(" + ",".join(exprs) + ")"


# 隐藏辅助版保留作备份，正式采用固定列顺序的合并公式。
def _q(value: str) -> str:
    return f'"{value}"'


def priority_order_expr(cols: list[str], stat_row: int = 5) -> str:
    if len(cols) != 3:
        raise ValueError("priority_order_expr supports exactly 3 strategy columns")
    a, b, c = cols
    ea = f"ABS(${a}${stat_row})"
    eb = f"ABS(${b}${stat_row})"
    ec = f"ABS(${c}${stat_row})"
    pair_bc = f'IF({eb}>={ec},{_q(f"{b},{c}")},{_q(f"{c},{b}")})'
    pair_ac = f'IF({ea}>={ec},{_q(f"{a},{c}")},{_q(f"{c},{a}")})'
    pair_ab = f'IF({ea}>={eb},{_q(f"{a},{b}")},{_q(f"{b},{a}")})'
    return (
        f'IF({ea}>=MAX({eb},{ec}),{_q(a)}&","&{pair_bc},'
        f'IF({eb}>=MAX({ea},{ec}),{_q(b)}&","&{pair_ac},'
        f'{_q(c)}&","&{pair_ab}))'
    )


def priority_merge_expr(cols: list[str], helper_col: str, row: int) -> str:
    n = len(cols)
    refs = []
    for k in range(1, n + 1):
        start = (k - 1) * 10 + 1
        token = (
            f'TRIM(MID(SUBSTITUTE(${helper_col}$12,",",REPT(" ",10)),'
            f"{start},10))"
        )
        refs.append(f"INDIRECT({token}&ROW())")
    expr = refs[-1]
    for ref in reversed(refs[:-1]):
        expr = f"IF({ref}<>\"\",{ref},{expr})"
    return expr


def etf_label(tokens: list[str]) -> str:
    etf = tokens[0]
    suffix = "_".join(tokens[1:])
    if suffix == "up":
        return f"{etf}>0"
    if suffix == "down":
        return f"{etf}<0"
    if suffix == "big_up":
        return f"{etf}>=1%"
    if suffix == "big_down":
        return f"{etf}<=-1%"
    if suffix == "gt2":
        return f"{etf}>2%"
    if suffix == "lt-2":
        return f"{etf}<-2%"
    return f"{etf} {suffix}"


def ext_label(part: str) -> str:
    import scan_v4_conditions as sc
    short = {
        "VIX_Close": "VIX收盘",
        "VIX_Chg%": "VIX",
        "VIX_5dChg": "VIX5d",
        "VIX_20dVol": "VIX20d",
        "TNX_Yield": "TNX",
        "TNX_ChgBp": "TNXbp",
        "CreditSpread": "信用利差",
        "JNKSpread": "高收益利差",
        "StkBonCorr": "股债相关",
        "USDGoldRatio": "美元/黄金",
        "SectRotation": "板块轮动",
        "VIX_TNX_Ratio": "VIX/TNX",
        "YldCurveProxy": "收益率曲线",
    }
    for safe, col_name in sc.EXTERNAL_COLS.items():
        prefix = f"ext_{safe}_"
        if part.startswith(prefix):
            op = part[len(prefix):]
            name = short.get(col_name, col_name)
            if op == "up":
                return f"{name}>0"
            if op == "down":
                return f"{name}<0"
            threshold = float(op[2:].replace("_", "."))
            if op.startswith("ge"):
                return f"{name}>={threshold}%"
            return f"{name}<={threshold}%"
    return part


def self_label(suffix: str) -> str:
    if suffix == "up":
        return "该基金当日回报>0"
    if suffix == "down":
        return "该基金当日回报<0"
    if suffix == "big_up":
        return "该基金当日回报>=2%"
    if suffix == "big_down":
        return "该基金当日回报<=-2%"
    if suffix == "3up":
        return "该基金连续3日上涨"
    if suffix == "3down":
        return "该基金连续3日下跌"
    if suffix == "5up":
        return "该基金连续5日上涨"
    if suffix == "5down":
        return "该基金连续5日下跌"
    return suffix


def condition_label(condition: str) -> str:
    if condition.startswith("combo_"):
        parts = condition[len("combo_"):].split("__")
        labels = []
        for p in parts:
            if p.startswith("ext_"):
                labels.append(ext_label(p))
            elif p.startswith("self_"):
                labels.append(self_label(p.removeprefix("self_")))
            else:
                labels.append(etf_label(p.split("_")))
        return " 且 ".join(labels) + "，买基金第二天"
    if condition.startswith("ext_"):
        return ext_label(condition) + "，买基金第二天"
    if condition.startswith("self_"):
        return self_label(condition.removeprefix("self_")) + "，买基金第二天"
    tokens = condition.split("_")
    if len(tokens) == 6:
        labels = [etf_label([tokens[0], tokens[1]]), etf_label([tokens[2], tokens[3]]), etf_label([tokens[4], tokens[5]])]
        return " 且 ".join(labels) + "，买基金第二天"
    if len(tokens) == 4:
        labels = [etf_label([tokens[0], tokens[1]]), etf_label([tokens[2], tokens[3]])]
        return " 且 ".join(labels) + "，买基金第二天"
    return etf_label(tokens) + "，买基金第二天"


def load_fund_adj(ticker: str) -> dict:
    p = config.RESULT_ROOT / "最新成果" / "数据" / "数据_原始" / "profunds" / "adj_close_nasdaq" / f"{ticker}.json"
    data = json.loads(p.read_text(encoding="utf-8-sig"))
    rows = data["data"]["tradesTable"]["rows"]
    out = {}
    for r in rows:
        d = pd.to_datetime(r["date"], format="%m/%d/%Y")
        out[d] = float(r["adjustedClose"])
    return out


def load_etf_ohlc(ticker: str) -> dict:
    if ticker == "XLC":
        p = config.RESULT_ROOT / "最新成果" / "数据" / "数据_原始" / "etfs_extended" / "XLC.csv"
    else:
        p = config.RESULT_ROOT / "最新成果" / "数据" / "数据_原始" / "etfs" / f"{ticker}_historical.csv"
    df = pd.read_csv(p)
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y")
    out = {}
    for rec in df.to_dict("records"):
        out[rec["Date"]] = {
            "open": float(rec["Open"]),
            "high": float(rec["High"]),
            "low": float(rec["Low"]),
            "close": float(rec["Close"]),
            "adj": float(rec["Adj Close"]),
        }
    return out


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    master = s4.load_master()
    dates_all = sorted(pd.to_datetime(master["Date"]), reverse=True)[:WINDOW_DAYS]
    dates_all = sorted(dates_all, reverse=True)

    mapping = pd.read_csv(config.V4_OUT / "v4_strategy_mapping_m30_v2.csv", keep_default_na=False)
    strat = {t: mapping[mapping["ticker"] == t].copy() for t in FUNDS}
    for t in FUNDS:
        strat[t]["full_avg"] = pd.to_numeric(strat[t]["full_avg"], errors="coerce")
        strat[t] = strat[t].sort_values("full_avg", key=lambda s: s.abs(), ascending=False)
    name_map = pd.read_csv(config.MIDDLE / "通用" / "文件" / "基金名称映射.csv", keep_default_na=False)
    name_by_ticker = dict(zip(name_map["ticker"], name_map["name"]))
    fund_label = {t: f"{name_by_ticker.get(t, t)}（{t}）" for t in FUNDS}

    base_headers = []
    for t in FUNDS:
        base_headers += [f"{fund_label[t]} Adj Close"]
    for t in FUNDS:
        base_headers += [f"{fund_label[t]}回报"]
    for e in ETF_ALL:
        base_headers += [f"{e} Open", f"{e} High", f"{e} Low", f"{e} Close", f"{e} Adj Close"]
    for e in ETF_ALL:
        base_headers += [f"{e}回报"]
    for e in EXT_RAW_COLS:
        base_headers += [e]
    base_headers += ["VIX回报"]
    base_headers += EXT_COLS_USED

    col_idx = 2
    fund_price_col = {}
    fund_return_col = {}
    for t in FUNDS:
        fund_price_col[t] = get_column_letter(col_idx); col_idx += 1
    for t in FUNDS:
        fund_return_col[t] = get_column_letter(col_idx); col_idx += 1
    etf_open_col = {}
    etf_high_col = {}
    etf_low_col = {}
    etf_close_col = {}
    etf_adj_col = {}
    etf_return_col = {}
    for e in ETF_ALL:
        etf_open_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_high_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_low_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_close_col[e] = get_column_letter(col_idx); col_idx += 1
        etf_adj_col[e] = get_column_letter(col_idx); col_idx += 1
    for e in ETF_ALL:
        etf_return_col[e] = get_column_letter(col_idx); col_idx += 1
    ext_price_col = {}
    ext_return_col = {}
    for e in EXT_RAW_COLS:
        ext_price_col[e] = get_column_letter(col_idx); col_idx += 1
    for e in EXT_RAW_COLS:
        ext_return_col[e] = get_column_letter(col_idx); col_idx += 1
    ext_col = {}
    for e in EXT_COLS_USED:
        ext_col[e] = get_column_letter(col_idx); col_idx += 1
    colmap = {**fund_return_col, **etf_return_col, **ext_col}

    per_fund = {t: list(strat[t]["condition"]) for t in FUNDS}
    headers = ["日期"] + base_headers
    for t in FUNDS:
        for c in per_fund[t]:
            headers.append(condition_label(c))
        headers.append(f"{fund_label[t]}合并效果")
    n_cols = len(headers)

    wb = Workbook()
    ws_note = wb.active
    ws_note.title = "说明"
    ws_note["A1"] = "两基金 m30 新版式示例"
    ws_note["A1"].font = title_font
    for i, line in enumerate(
        [
            "公式说明：",
            "1. 统计头公式（第2-10行）",
            "   Hit Ratio = Up Count / (Up Count + Down Count)；",
            "   Up Count = COUNTIF(该列数据区, \">0\")；Down Count = COUNTIF(该列数据区, \"<0\")；",
            "   Average = AVERAGE(该列数据区)；Max/Min/Count/Std/Sum 对应 MAX/MIN/COUNT/STDEV/SUM。",
            "2. 回报率公式",
            "   =IF(COUNT(B15:B16)=2,ROUND((B15/B16-1)*100,4),\"\")",
            "   B15 是今日 Adj Close，B16 是前一交易日 Adj Close；",
            "   COUNT 判断两天价格都存在才计算当日回报率，否则留空，避免 #DIV/0!。",
            "   VIX_Close 的回报率公式相同；VIX_Chg% 为外部涨跌幅列，可与计算值核对。",
            "   ETF 原始数据列为 Open/High/Low/Close/Adj Close，回报率由 Adj Close 计算；",
            "   表内按区块排列：基金原始价→基金回报；ETF 原始 OHLC→ETF 回报；外部 VIX→回报。",
            "   所有回报率均以百分比显示并保留4位小数（公式 ROUND 到4位，显示格式 0.0000%）。",
            "3. 策略公式",
            "   =IF(条件, 该基金次日实际回报, \"\")",
            "   条件引用回报率列和阈值参数；满足条件时显示次日实际回报（日期降序，次日=上一行）。",
            "4. 合并公式",
            "   策略列按 |全历史Average| 从高到低排列，合并列取第一个非空策略：",
            "   =IF(策略1<>\"\",策略1,IF(策略2<>\"\",策略2,策略3))",
            "   同一天多条策略触发时，历史 |Average| 最大的策略生效；全部未触发则留空。",
            "5. 阈值参数",
            "   每个策略列正上方的第11/12/13行是该策略条件阈值：二条件用12/13行，三条件额外用11行；",
            "   修改数字可调整条件，公式自动更新。",
        ],
        start=3,
    ):
        ws_note.cell(row=i, column=1, value=line)
    ws_note.column_dimensions["A"].width = 110

    ws = wb.create_sheet("数据")
    ws.append([])
    for i, label in enumerate(STAT_NAMES, start=2):
        ws.cell(row=i, column=1, value=label)
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(headers)
    ds = 15
    de = ds + len(dates_all) - 1

    for c_idx in range(2, n_cols + 1):
        col = get_column_letter(c_idx)
        ws[f"{col}2"] = f"={col}3/({col}3+{col}4)"
        ws[f"{col}3"] = f'=COUNTIF({col}{ds}:{col}{de},">0")'
        ws[f"{col}4"] = f'=COUNTIF({col}{ds}:{col}{de},"<0")'
        ws[f"{col}5"] = f"=AVERAGE({col}{ds}:{col}{de})"
        ws[f"{col}6"] = f"=MAX({col}{ds}:{col}{de})"
        ws[f"{col}7"] = f"=MIN({col}{ds}:{col}{de})"
        ws[f"{col}8"] = f"=COUNT({col}{ds}:{col}{de})"
        ws[f"{col}9"] = f"=STDEV({col}{ds}:{col}{de})"
        ws[f"{col}10"] = f"=SUM({col}{ds}:{col}{de})"

    fund_adj = {t: load_fund_adj(t) for t in FUNDS}
    etf_ohlc = {e: load_etf_ohlc(e) for e in ETF_ALL}
    ext_val = {e: dict(zip(pd.to_datetime(master["Date"]), master[e])) for e in EXT_RAW_COLS + EXT_COLS_USED}

    merged_idx = {}
    cur = 2 + len(base_headers)
    for t in FUNDS:
        cur += len(per_fund[t])
        merged_idx[t] = cur
        cur += 1
    strat_cols = {}
    cur = 2 + len(base_headers)
    for t in FUNDS:
        cols = []
        for _ in per_fund[t]:
            cols.append(cur)
            cur += 1
        strat_cols[t] = cols
        cur += 1

    strategy_thr_refs: dict[int, list[str]] = {}
    for t in FUNDS:
        for c_idx, condition in zip(strat_cols[t], per_fund[t]):
            parts = condition_parts(condition, colmap)
            values = [part_threshold(p) for p in parts]
            col = get_column_letter(c_idx)
            ws.cell(row=12, column=c_idx, value=values[0])
            ws.cell(row=13, column=c_idx, value=values[1] if len(values) > 1 else "")
            if len(values) > 2:
                ws.cell(row=11, column=c_idx, value=values[2])
            for r in (11, 12, 13):
                ws.cell(row=r, column=c_idx).font = body_font
            strategy_thr_refs[c_idx] = [f"${col}$12", f"${col}$13"]
            if len(values) > 2:
                strategy_thr_refs[c_idx].append(f"${col}$11")

    price_cols = [fund_price_col[t] for t in FUNDS] + [etf_adj_col[e] for e in ETF_ALL] + [ext_price_col[e] for e in EXT_RAW_COLS]
    ret_cols = [fund_return_col[t] for t in FUNDS] + [etf_return_col[e] for e in ETF_ALL] + [ext_return_col[e] for e in EXT_RAW_COLS]

    for i, d in enumerate(dates_all):
        r = ds + i
        row = [d.strftime("%Y/%m/%d")]
        for t in FUNDS:
            row.append(round(fund_adj[t].get(d, float("nan")), 4))
        for t in FUNDS:
            row.append("")
        for e in ETF_ALL:
            rec = etf_ohlc[e].get(d, {})
            row.append(round(rec.get("open", float("nan")), 4))
            row.append(round(rec.get("high", float("nan")), 4))
            row.append(round(rec.get("low", float("nan")), 4))
            row.append(round(rec.get("close", float("nan")), 4))
            row.append(round(rec.get("adj", float("nan")), 4))
        for e in ETF_ALL:
            row.append("")
        for e in EXT_RAW_COLS:
            row.append(round(ext_val[e].get(d, float("nan")), 4))
        for e in EXT_RAW_COLS:
            row.append("")
        for e in EXT_COLS_USED:
            row.append(round(ext_val[e].get(d, float("nan")), 4))
        ws.append(row)
        for pcol, rcol in zip(price_cols, ret_cols):
            if i < len(dates_all) - 1:
                ws[f"{rcol}{r}"] = (
                    f"=IF(COUNT({pcol}{r}:{pcol}{r + 1})=2,"
                    f"ROUND(({pcol}{r}/{pcol}{r + 1}-1)*100,4),\"\")"
                )
            else:
                ws[f"{rcol}{r}"] = ""
        for t in FUNDS:
            fund_col = colmap[t]
            for c_idx, condition in zip(strat_cols[t], per_fund[t]):
                if i == 0:
                    ws.cell(row=r, column=c_idx, value="")
                else:
                    cond = condition_expr(condition, t, r, colmap, fund_col, strategy_thr_refs[c_idx])
                    ws.cell(row=r, column=c_idx, value=f"=IF({cond},{fund_col}{r - 1},\"\")")
            cols = [get_column_letter(c) for c in strat_cols[t]]
            expr = f"{cols[-1]}{r}"
            for col in reversed(cols[:-1]):
                expr = f"IF({col}{r}<>\"\",{col}{r},{expr})"
            ws.cell(row=r, column=merged_idx[t], value=f"={expr}")

    pct_cols = (
        list(fund_return_col.values())
        + list(etf_return_col.values())
        + list(ext_return_col.values())
        + list(ext_col.values())
    )
    for t in FUNDS:
        pct_cols += [get_column_letter(c) for c in strat_cols[t]]
        pct_cols.append(get_column_letter(merged_idx[t]))
    for r in range(ds, de + 1):
        for col in pct_cols:
            ws[f"{col}{r}"].number_format = '0.0000"%"'

    for r in range(2, 11):
        ws.cell(row=r, column=1).font = head_font
        for c in range(2, n_cols + 1):
            ws.cell(row=r, column=c).font = body_font
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=14, column=c)
        cell.font = head_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(ds, de + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = body_font
    ws.column_dimensions["A"].width = 12
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "B15"

    out_dir = config.RESULT_ROOT / "示例审批"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "两基金_m30新版式示例_公式说明版.xlsx"
    if os.environ.get("SAMPLE_OUT"):
        out_path = pathlib.Path(os.environ["SAMPLE_OUT"])
    wb.save(out_path)
    print("saved:", out_path)
    print("cols:", n_cols, "data rows:", len(dates_all))


if __name__ == "__main__":
    main()
